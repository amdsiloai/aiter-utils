#!/usr/bin/env python3
"""
Benchmark Analysis Tool

Fetches benchmark results from Azure Databricks MLflow and generates
comparative analysis reports (CSV + Excel).

Usage:
    # Process specific models
    python benchmark_analysis.py --models "Llama-3.3-70B" "DeepSeek-R1" "Qwen3-235B"
    
    # Process all models
    python benchmark_analysis.py --models "*"
    
    # With custom output directory
    python benchmark_analysis.py --models "Llama-3.3-70B" --output-dir ./results --run-pattern "oob_"

Environment Variables:
    DATABRICKS_WORKSPACE_URL: Azure Databricks workspace URL
    DATABRICKS_API_KEY: Databricks personal access token
"""

import os
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# Load .env file if present
try:
    from dotenv import load_dotenv
    env_path = Path(__file__).parent / '.env'
    if env_path.exists():
        load_dotenv(env_path)
    else:
        load_dotenv()
except ImportError:
    pass

import math
import re
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import from utils module
from utils import (
    DatabricksMLflowClient,
    process_all_experiments,
    load_model_results,
    select_latest_version,
    group_results_by_config,
    merge_configurations,
    calculate_comparisons,
    calculate_geomean_speedup,
    _create_short_sheet_name,
    extract_framework_version,
)


# =============================================================================
# Version Change Analysis Functions (NEW)
# =============================================================================

def calculate_version_changes(
    results: Dict[tuple, pd.DataFrame],
    metric: str = 'total_token_throughput'
) -> List[Dict]:
    """
    Calculate performance changes between different versions of the same framework.
    
    Args:
        results: Dictionary mapping (gpu, framework, version, precision, parallelism) -> DataFrame
        metric: Metric to use for comparison (default: total_token_throughput)
    
    Returns:
        List of dictionaries with version change information
    """
    grouped = {}
    for (gpu, framework, version, precision, parallelism), df in results.items():
        key = (gpu, framework, precision, parallelism)
        if key not in grouped:
            grouped[key] = []
        
        if metric in df.columns:
            values = df[metric].replace([np.inf, -np.inf], np.nan).dropna()
            values = values[values > 0]
            if len(values) > 0:
                geomean = np.exp(np.log(values).mean())
            else:
                geomean = np.nan
        else:
            geomean = np.nan
        
        docker_image = 'unknown'
        if 'docker_image' in df.columns and len(df) > 0:
            docker_image = df['docker_image'].iloc[0]
        
        model_name = 'unknown'
        if 'model' in df.columns and len(df) > 0:
            model_name = str(df['model'].iloc[0])
        elif 'model_name' in df.columns and len(df) > 0:
            model_name = str(df['model_name'].iloc[0])
        
        grouped[key].append({
            'version': version,
            'docker_image': docker_image,
            'geomean': geomean,
            'model': model_name,
            'num_configs': len(df)
        })
    
    version_changes = []
    
    for (gpu, framework, precision, parallelism), versions in grouped.items():
        if len(versions) < 2:
            continue
        
        versions_sorted = sorted(versions, key=lambda x: x['version'])
        
        for i in range(1, len(versions_sorted)):
            old_ver = versions_sorted[i-1]
            new_ver = versions_sorted[i]
            
            if np.isnan(old_ver['geomean']) or np.isnan(new_ver['geomean']):
                continue
            
            change_pct = ((new_ver['geomean'] - old_ver['geomean']) / old_ver['geomean']) * 100
            
            version_changes.append({
                'Model': new_ver['model'],
                'GPU': gpu,
                'Framework': framework.upper(),
                'Precision': precision,
                'Parallelism': parallelism,
                'Old Version': old_ver['version'],
                'New Version': new_ver['version'],
                'Old Docker': old_ver['docker_image'],
                'New Docker': new_ver['docker_image'],
                f'Old Geomean ({metric})': old_ver['geomean'],
                f'New Geomean ({metric})': new_ver['geomean'],
                'Change (%)': change_pct,
                'Old Configs': old_ver['num_configs'],
                'New Configs': new_ver['num_configs'],
            })
    
    return version_changes


def calculate_competitive_changes(
    results: Dict[tuple, pd.DataFrame],
    gpu1: str = 'MI355X',
    gpu2: str = 'B200',
    metric: str = 'total_token_throughput'
) -> List[Dict]:
    """
    Calculate how the competitive ratio (gpu1/gpu2) changes between framework versions.
    Uses the latest gpu2 (B200) result as a fixed reference baseline.
    
    Now calculates ratios using only MATCHING configurations (same input_len, output_len, concurrency)
    to be consistent with the Summary sheet calculation.
    """
    gpu_framework_data = {}
    for (gpu, framework, version, precision, parallelism), df in results.items():
        key = (gpu, framework, precision, parallelism)
        if key not in gpu_framework_data:
            gpu_framework_data[key] = []
        
        docker_image = 'unknown'
        if 'docker_image' in df.columns and len(df) > 0:
            docker_image = df['docker_image'].iloc[0]
        
        model_name = 'unknown'
        if 'model' in df.columns and len(df) > 0:
            model_name = str(df['model'].iloc[0])
        elif 'model_name' in df.columns and len(df) > 0:
            model_name = str(df['model_name'].iloc[0])
        
        # Store the full DataFrame for later matching
        gpu_framework_data[key].append({
            'version': version,
            'df': df,
            'docker_image': docker_image,
            'model': model_name
        })
    
    competitive_changes = []
    
    framework_configs = set()
    for (gpu, framework, precision, parallelism) in gpu_framework_data.keys():
        framework_configs.add((framework, precision, parallelism))
    
    def calculate_matched_ratio(df1, df2, metric):
        """Calculate geomean of ratios for matching configurations only."""
        # Normalize column names for matching
        config_cols = ['random_input_len', 'random_output_len', 'max_concurrency']
        
        # Check if required columns exist
        for col in config_cols:
            if col not in df1.columns or col not in df2.columns:
                return np.nan, 0
        
        if metric not in df1.columns or metric not in df2.columns:
            return np.nan, 0
        
        # Create config keys for matching
        df1_clean = df1[config_cols + [metric]].dropna()
        df2_clean = df2[config_cols + [metric]].dropna()
        
        # Merge on config columns
        merged = pd.merge(
            df1_clean, df2_clean,
            on=config_cols,
            suffixes=('_gpu1', '_gpu2')
        )
        
        if len(merged) == 0:
            return np.nan, 0
        
        # Calculate ratios for each matched config
        gpu1_vals = merged[f'{metric}_gpu1']
        gpu2_vals = merged[f'{metric}_gpu2']
        
        # Avoid division by zero
        valid_mask = (gpu2_vals > 0) & (gpu1_vals > 0)
        if valid_mask.sum() == 0:
            return np.nan, 0
        
        ratios = gpu1_vals[valid_mask] / gpu2_vals[valid_mask]
        
        # Calculate geomean of ratios
        geomean_ratio = np.exp(np.log(ratios).mean())
        return geomean_ratio, len(ratios)
    
    for (framework, precision, parallelism) in framework_configs:
        gpu1_key = (gpu1, framework, precision, parallelism)
        gpu2_key = (gpu2, framework, precision, parallelism)
        
        if gpu1_key not in gpu_framework_data or gpu2_key not in gpu_framework_data:
            continue
        
        gpu1_versions = gpu_framework_data[gpu1_key]
        gpu2_versions = gpu_framework_data[gpu2_key]
        
        # Get latest B200 version as reference
        gpu2_sorted = sorted(gpu2_versions, key=lambda x: x['version'], reverse=True)
        gpu2_reference = gpu2_sorted[0]
        
        # Sort MI355X versions and filter valid ones
        gpu1_sorted = sorted(gpu1_versions, key=lambda x: x['version'])
        
        # Calculate matched ratios for each MI355X version against B200 reference
        gpu1_with_ratios = []
        for v in gpu1_sorted:
            ratio, n_configs = calculate_matched_ratio(v['df'], gpu2_reference['df'], metric)
            if not np.isnan(ratio) and n_configs > 0:
                v['matched_ratio'] = ratio
                v['n_matched_configs'] = n_configs
                gpu1_with_ratios.append(v)
        
        if len(gpu1_with_ratios) < 2:
            continue
        
        for i in range(1, len(gpu1_with_ratios)):
            old_ver = gpu1_with_ratios[i-1]
            new_ver = gpu1_with_ratios[i]
            
            old_ratio = old_ver['matched_ratio']
            new_ratio = new_ver['matched_ratio']
            
            ratio_change = new_ratio - old_ratio
            ratio_change_pct = ((new_ratio - old_ratio) / old_ratio) * 100 if old_ratio > 0 else 0
            
            competitive_changes.append({
                'Model': new_ver['model'],
                'Framework': framework.upper(),
                'Precision': precision,
                'Parallelism': parallelism,
                f'{gpu1} Old Version': old_ver['version'],
                f'{gpu1} New Version': new_ver['version'],
                f'{gpu2} Reference Version': gpu2_reference['version'],
                f'Old {gpu1}/{gpu2} Ratio': old_ratio,
                f'New {gpu1}/{gpu2} Ratio': new_ratio,
                'Ratio Change': ratio_change,
                'Ratio Change (%)': ratio_change_pct,
                f'Matched Configs': new_ver['n_matched_configs'],
                f'{gpu1} Old Docker': old_ver['docker_image'],
                f'{gpu1} New Docker': new_ver['docker_image'],
                f'{gpu2} Reference Docker': gpu2_reference['docker_image'],
            })
    
    return competitive_changes


# =============================================================================
# Enhanced Geomean Summary (Fixed for Excel)
# =============================================================================

def add_geomean_summary(df: pd.DataFrame, gpu1: str = 'MI355X', gpu2: str = 'B200') -> pd.DataFrame:
    """Add summary rows with geomean - fixed to avoid Excel formula interpretation"""
    comparison_cols = ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]
    existing_comparison_cols = [col for col in comparison_cols if col in df.columns]

    if not existing_comparison_cols:
        return df

    summary_rows = []
    
    separator = {col: None for col in df.columns}
    summary_rows.append(separator)
    
    # Use --- instead of === to avoid Excel formula interpretation
    header = {col: None for col in df.columns}
    header['random_input_len'] = '--- GEOMEAN SUMMARY ---'
    summary_rows.append(header)
    
    geomean_row = {col: None for col in df.columns}
    geomean_row['random_input_len'] = 'Geomean'

    for comp_col in existing_comparison_cols:
        geomean = calculate_geomean_speedup(df, comp_col)
        geomean_row[comp_col] = geomean if not np.isnan(geomean) else None

    summary_rows.append(geomean_row)
    
    summary_df = pd.DataFrame(summary_rows)
    df_summary = pd.concat([df, summary_df], ignore_index=True)

    return df_summary


# =============================================================================
# Excel Formatting Functions (Enhanced)
# =============================================================================

def format_sheet_with_merged_headers(worksheet, df: pd.DataFrame, gpu1: str = 'MI355X', gpu2: str = 'B200'):
    """Format Excel sheet with merged headers - with fixes for data type handling"""
    metric_columns = ['median_itl_ms', 'median_ttft_ms', 'median_tpot_ms',
                     'median_e2el_ms', 'output_throughput', 'total_token_throughput']

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    missing_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

    worksheet.insert_rows(1)

    def extract_framework_header(col_name: str) -> str:
        for metric in metric_columns:
            if col_name.endswith(f'_{metric}'):
                return col_name[:-len(f'_{metric}')]
        return col_name

    framework_to_cols = {}
    for col_idx, col_name in enumerate(df.columns, 1):
        is_metric_col = any(col_name.endswith(f'_{metric}') for metric in metric_columns)
        if is_metric_col:
            framework_header = extract_framework_header(col_name)
            if framework_header not in framework_to_cols:
                framework_to_cols[framework_header] = []
            framework_to_cols[framework_header].append(col_idx)

    framework_groups = []
    comparison_groups = []
    col_idx = 1

    for col_name in df.columns:
        if col_name in ['random_input_len', 'random_output_len', 'max_concurrency']:
            col_idx += 1
            continue

        comparison_cols = ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]
        if col_name in comparison_cols:
            if not comparison_groups or comparison_groups[-1]['end_col'] != col_idx - 1:
                comparison_groups.append({'start_col': col_idx, 'end_col': col_idx})
            else:
                comparison_groups[-1]['end_col'] = col_idx
            col_idx += 1
            continue

        if 'best framework' in col_name:
            col_idx += 1
            continue

        is_metric_col = any(col_name.endswith(f'_{metric}') for metric in metric_columns)

        if is_metric_col:
            framework_header = extract_framework_header(col_name)
            if not framework_groups or framework_groups[-1]['header'] != framework_header:
                framework_groups.append({'header': framework_header, 'start_col': col_idx, 'end_col': col_idx})
            else:
                framework_groups[-1]['end_col'] = col_idx

        col_idx += 1

    col_idx = 1
    comparison_col_indices = []
    metric_col_indices = []

    for col_name in df.columns:
        col_letter = get_column_letter(col_idx)

        if col_name in ['random_input_len', 'random_output_len', 'max_concurrency']:
            worksheet.merge_cells(f'{col_letter}1:{col_letter}2')
            cell = worksheet[f'{col_letter}1']
            cell.value = col_name
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, size=9)
            cell.fill = header_fill
            cell.border = thin_border

        elif 'best framework' in col_name:
            worksheet.merge_cells(f'{col_letter}1:{col_letter}2')
            cell = worksheet[f'{col_letter}1']
            cell.value = col_name
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, size=8)
            cell.fill = header_fill
            cell.border = thin_border

        elif col_name in ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]:
            cell_row2 = worksheet.cell(row=2, column=col_idx)
            cell_row2.value = col_name
            cell_row2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_row2.font = Font(bold=True, size=8)
            cell_row2.fill = header_fill
            cell_row2.border = thin_border
            comparison_col_indices.append(col_idx)

        else:
            cell_row2 = worksheet.cell(row=2, column=col_idx)
            for metric in metric_columns:
                if col_name.endswith(f'_{metric}'):
                    cell_row2.value = metric
                    metric_col_indices.append(col_idx)
                    break
            cell_row2.alignment = Alignment(horizontal='center', vertical='center')
            cell_row2.font = Font(bold=True, size=8)
            cell_row2.fill = header_fill
            cell_row2.border = thin_border

        col_idx += 1

    for group in framework_groups:
        start_col_letter = get_column_letter(group['start_col'])
        end_col_letter = get_column_letter(group['end_col'])

        if group['start_col'] != group['end_col']:
            worksheet.merge_cells(f'{start_col_letter}1:{end_col_letter}1')

        cell = worksheet[f'{start_col_letter}1']
        cell.value = group['header']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.border = thin_border

        for col in range(group['start_col'], group['end_col'] + 1):
            worksheet.cell(row=1, column=col).border = thin_border
            worksheet.cell(row=1, column=col).fill = header_fill

    for group in comparison_groups:
        start_col_letter = get_column_letter(group['start_col'])
        end_col_letter = get_column_letter(group['end_col'])

        if group['start_col'] != group['end_col']:
            worksheet.merge_cells(f'{start_col_letter}1:{end_col_letter}1')

        cell = worksheet[f'{start_col_letter}1']
        cell.value = f"{gpu1} vs {gpu2}"
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.border = thin_border

        for col in range(group['start_col'], group['end_col'] + 1):
            worksheet.cell(row=1, column=col).border = thin_border
            worksheet.cell(row=1, column=col).fill = header_fill

    def is_missing_or_zero(val):
        if val is None or val == '':
            return True
        if isinstance(val, (int, float)):
            try:
                return math.isnan(val) or val == 0
            except (TypeError, ValueError):
                return True
        return False
    
    def is_valid_number(val):
        if val is None or val == '':
            return False
        if isinstance(val, (int, float)):
            try:
                return not math.isnan(val)
            except (TypeError, ValueError):
                return False
        return False

    for row_idx in range(3, worksheet.max_row + 1):
        first_cell_value = worksheet.cell(row=row_idx, column=1).value
        is_summary_row = first_cell_value in ['--- GEOMEAN SUMMARY ---', 'Geomean', '', None]
        is_geomean_row = first_cell_value == 'Geomean'

        framework_missing = {}
        for fw_header, fw_cols in framework_to_cols.items():
            all_zero_or_missing = True
            for fw_col_idx in fw_cols:
                cell_val = worksheet.cell(row=row_idx, column=fw_col_idx).value
                if cell_val is not None and cell_val != '' and cell_val != 0:
                    try:
                        if float(cell_val) != 0:
                            all_zero_or_missing = False
                            break
                    except (ValueError, TypeError):
                        all_zero_or_missing = False
                        break
            framework_missing[fw_header] = all_zero_or_missing and not is_summary_row

        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if is_geomean_row:
                cell.font = Font(bold=True)

            if col_idx in metric_col_indices and not is_summary_row:
                col_framework = None
                for fw_header, fw_cols in framework_to_cols.items():
                    if col_idx in fw_cols:
                        col_framework = fw_header
                        break

                if col_framework and framework_missing.get(col_framework, False):
                    cell.fill = missing_fill
                elif is_missing_or_zero(cell.value):
                    cell.fill = missing_fill

            if col_idx in comparison_col_indices:
                if not is_summary_row:
                    if not is_valid_number(cell.value):
                        cell.fill = missing_fill
                    else:
                        cell.number_format = '0.0%'
                elif is_geomean_row and is_valid_number(cell.value):
                    cell.number_format = '0.0%'

    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        col_name = df.columns[col_idx - 1]

        if col_name in ['random_input_len', 'random_output_len', 'max_concurrency']:
            worksheet.column_dimensions[col_letter].width = 10
        elif 'vs' in col_name or 'best framework' in col_name:
            worksheet.column_dimensions[col_letter].width = 12
        else:
            worksheet.column_dimensions[col_letter].width = 12

    worksheet.freeze_panes = 'D3'


def format_summary_sheet(worksheet, df: pd.DataFrame, gpu1: str = 'MI355X', gpu2: str = 'B200'):
    """Format summary sheet with merged headers and heatmap coloring"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    worksheet.insert_rows(1)

    comparison_cols = ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]
    comparison_start = None
    comparison_end = None
    comparison_col_indices = []

    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name in comparison_cols:
            if comparison_start is None:
                comparison_start = col_idx
            comparison_end = col_idx
            comparison_col_indices.append(col_idx)

    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)

        if col_name in comparison_cols:
            cell_row2 = worksheet.cell(row=2, column=col_idx)
            cell_row2.value = col_name
            cell_row2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_row2.font = Font(bold=True, size=8)
            cell_row2.fill = header_fill
            cell_row2.border = thin_border
        else:
            worksheet.merge_cells(f'{col_letter}1:{col_letter}2')
            cell = worksheet[f'{col_letter}1']
            cell.value = col_name
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, size=9)
            cell.fill = header_fill
            cell.border = thin_border

    if comparison_start and comparison_end:
        start_letter = get_column_letter(comparison_start)
        end_letter = get_column_letter(comparison_end)
        worksheet.merge_cells(f'{start_letter}1:{end_letter}1')

        cell = worksheet[f'{start_letter}1']
        cell.value = f"{gpu1} vs {gpu2}"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.border = thin_border

        for col_idx in range(comparison_start, comparison_end + 1):
            worksheet.cell(row=1, column=col_idx).border = thin_border
            worksheet.cell(row=1, column=col_idx).fill = header_fill

    def is_valid_number(val):
        if val is None or val == '':
            return False
        if isinstance(val, (int, float)):
            try:
                return not math.isnan(val)
            except (TypeError, ValueError):
                return False
        return False

    for row_idx in range(3, worksheet.max_row + 1):
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if col_idx in comparison_col_indices:
                if is_valid_number(cell.value):
                    cell.number_format = '0.0%'
                    value = cell.value

                    if value >= 1.0:
                        cell.fill = green_fill
                    elif value >= 0.85:
                        cell.fill = light_green_fill
                    elif value >= 0.70:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill

    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        col_name = df.columns[col_idx - 1]

        if col_name == 'Model':
            worksheet.column_dimensions[col_letter].width = 25
        elif col_name in ['Precision', 'Parallelism']:
            worksheet.column_dimensions[col_letter].width = 10
        elif 'Best' in col_name:
            worksheet.column_dimensions[col_letter].width = 10
        elif 'Docker' in col_name:
            worksheet.column_dimensions[col_letter].width = 35
        else:
            worksheet.column_dimensions[col_letter].width = 11

    worksheet.freeze_panes = 'A3'


def format_version_changes_sheet(worksheet, df: pd.DataFrame):
    """Format version changes sheet with conditional coloring"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    change_col_idx = None
    for col_idx, col_name in enumerate(df.columns, 1):
        if 'Change (%)' in col_name:
            change_col_idx = col_idx
            break

    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            if col_idx == change_col_idx and cell.value is not None:
                try:
                    change_val = float(cell.value)
                    cell.number_format = '+0.0%;-0.0%;0.0%'
                    cell.value = change_val / 100
                    
                    if change_val >= 10:
                        cell.fill = green_fill
                        cell.font = Font(bold=True, color='006400')
                    elif change_val >= 2:
                        cell.fill = light_green_fill
                    elif change_val <= -10:
                        cell.fill = red_fill
                        cell.font = Font(bold=True, color='8B0000')
                    elif change_val <= -2:
                        cell.fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
                    else:
                        cell.fill = yellow_fill
                except (ValueError, TypeError):
                    pass

            col_name = df.columns[col_idx - 1]
            if 'Geomean' in col_name and cell.value is not None:
                try:
                    cell.number_format = '0.00'
                except (ValueError, TypeError):
                    pass

    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        col_name = df.columns[col_idx - 1]

        if col_name == 'Model':
            worksheet.column_dimensions[col_letter].width = 25
        elif 'Docker' in col_name:
            worksheet.column_dimensions[col_letter].width = 30
        elif 'Version' in col_name:
            worksheet.column_dimensions[col_letter].width = 12
        elif 'Geomean' in col_name:
            worksheet.column_dimensions[col_letter].width = 15
        elif col_name == 'Change (%)':
            worksheet.column_dimensions[col_letter].width = 12
        else:
            worksheet.column_dimensions[col_letter].width = 10

    worksheet.freeze_panes = 'A2'


def format_competitive_changes_sheet(worksheet, df: pd.DataFrame, gpu1: str = 'MI355X', gpu2: str = 'B200'):
    """Format competitive changes sheet"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    light_red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ratio_change_pct_col_idx = None
    ratio_col_indices = []
    
    for col_idx, col_name in enumerate(df.columns, 1):
        if 'Ratio Change (%)' in col_name:
            ratio_change_pct_col_idx = col_idx
        elif 'Ratio' in col_name and ('Old' in col_name or 'New' in col_name):
            ratio_col_indices.append(col_idx)

    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            col_name = df.columns[col_idx - 1]

            if col_idx in ratio_col_indices and cell.value is not None:
                try:
                    cell.number_format = '0.0%'
                except (ValueError, TypeError):
                    pass

            if col_idx == ratio_change_pct_col_idx and cell.value is not None:
                try:
                    change_val = float(cell.value)
                    cell.number_format = '+0.0%;-0.0%;0.0%'
                    cell.value = change_val / 100
                    
                    if change_val >= 10:
                        cell.fill = green_fill
                        cell.font = Font(bold=True, color='006400')
                    elif change_val >= 2:
                        cell.fill = light_green_fill
                    elif change_val <= -10:
                        cell.fill = red_fill
                        cell.font = Font(bold=True, color='8B0000')
                    elif change_val <= -2:
                        cell.fill = light_red_fill
                    else:
                        cell.fill = yellow_fill
                except (ValueError, TypeError):
                    pass

            if 'Geomean' in col_name and cell.value is not None:
                try:
                    cell.number_format = '0.00'
                except (ValueError, TypeError):
                    pass

    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        col_name = df.columns[col_idx - 1]

        if col_name == 'Model':
            worksheet.column_dimensions[col_letter].width = 25
        elif 'Version' in col_name:
            worksheet.column_dimensions[col_letter].width = 12
        elif 'Ratio' in col_name:
            worksheet.column_dimensions[col_letter].width = 14
        elif 'Geomean' in col_name:
            worksheet.column_dimensions[col_letter].width = 14
        else:
            worksheet.column_dimensions[col_letter].width = 10

    worksheet.freeze_panes = 'A2'


# =============================================================================
# Enhanced Comparative Analysis
# =============================================================================

def create_comparative_analysis_enhanced(
    parent_dir: str,
    output_file: str = "comparative_analysis.xlsx",
    gpu1: str = 'MI355X',
    gpu2: str = 'B200',
    verbose: bool = True
) -> Dict[str, pd.DataFrame]:
    """
    Create comparative analysis with enhanced features:
    - Version changes tracking
    - Competitive ratio changes
    - Docker image columns in summary
    - HuggingFace model names
    - Only include rows with comparative data
    - Recent additions tracking
    """
    parent_path = Path(parent_dir)

    if not parent_path.exists():
        raise ValueError(f"Directory not found: {parent_dir}")

    model_dirs = [d for d in parent_path.iterdir() if d.is_dir()]

    if verbose:
        print(f"Found {len(model_dirs)} model directories")
        print("=" * 80)

    all_results = {}
    all_warnings = []
    summary_data = []
    all_version_changes = []
    all_competitive_changes = []
    all_additions = []  # Track all benchmark files with timestamps
    
    # Get today's date for filtering recent additions
    today_str = datetime.now().strftime('%Y%m%d')

    for model_dir in model_dirs:
        model_name = model_dir.name

        if verbose:
            print(f"\nProcessing model: {model_name}")

        raw_results, additions = load_model_results(model_dir, track_additions=True)
        all_additions.extend(additions)

        if not raw_results:
            if verbose:
                print(f"  ⚠ No valid CSV files found")
            continue

        if verbose:
            print(f"  Found {len(raw_results)} result files")

        # Calculate version changes
        version_changes = calculate_version_changes(raw_results)
        all_version_changes.extend(version_changes)
        
        # Calculate competitive ratio changes
        competitive_changes = calculate_competitive_changes(raw_results, gpu1, gpu2)
        all_competitive_changes.extend(competitive_changes)
        
        if verbose and version_changes:
            print(f"  Found {len(version_changes)} version change(s)")
        if verbose and competitive_changes:
            print(f"  Found {len(competitive_changes)} competitive ratio change(s)")

        selected_results, warnings = select_latest_version(raw_results)

        if warnings:
            all_warnings.extend([f"{model_name}: {w}" for w in warnings])
            if verbose:
                for warning in warnings:
                    print(f"  ⚠ {warning}")

        config_groups = group_results_by_config(selected_results)

        if verbose:
            print(f"  Found {len(config_groups)} configuration(s): {list(config_groups.keys())}")

        for (precision, parallelism), config_results in config_groups.items():
            config_suffix = f"{parallelism}_{precision}"
            sheet_key = f"{model_name}_{config_suffix}"

            if verbose:
                print(f"\n  Processing config: {precision} / {parallelism}")

            merged_df = merge_configurations(config_results, gpu1=gpu1, gpu2=gpu2)

            if merged_df.empty:
                if verbose:
                    print(f"    ⚠ No data after merging")
                continue

            comparison_df = calculate_comparisons(merged_df, gpu1, gpu2)

            # Extract full HuggingFace model name
            hf_model_name = model_name
            for (gpu, framework), (version, docker_image, df) in config_results.items():
                if 'model' in df.columns and len(df) > 0:
                    full_model = df['model'].iloc[0]
                    if full_model and str(full_model) != 'nan' and pd.notna(full_model):
                        hf_model_name = str(full_model)
                        break

            summary_row = {
                'Model': hf_model_name,
                'Precision': precision,
                'Parallelism': parallelism,
            }

            if f"{gpu1} best framework" in comparison_df.columns:
                summary_row[f'{gpu1} Best'] = comparison_df[f"{gpu1} best framework"].iloc[0]
            if f"{gpu2} best framework" in comparison_df.columns:
                summary_row[f'{gpu2} Best'] = comparison_df[f"{gpu2} best framework"].iloc[0]

            # Extract docker images from column names
            def extract_docker_from_columns(df_cols, gpu, framework):
                for col in df_cols:
                    if gpu in col and framework.lower() in col.lower() and '(' in col and ')' in col:
                        start = col.find('(') + 1
                        end = col.find(')')
                        if start > 0 and end > start:
                            return col[start:end]
                return 'N/A'

            summary_row[f'{gpu1} vLLM Docker'] = extract_docker_from_columns(comparison_df.columns, gpu1, 'vllm')
            summary_row[f'{gpu1} SGLang Docker'] = extract_docker_from_columns(comparison_df.columns, gpu1, 'sglang')
            summary_row[f'{gpu1} ATOM Docker'] = extract_docker_from_columns(comparison_df.columns, gpu1, 'atom')
            summary_row[f'{gpu2} vLLM Docker'] = extract_docker_from_columns(comparison_df.columns, gpu2, 'vllm')
            summary_row[f'{gpu2} SGLang Docker'] = extract_docker_from_columns(comparison_df.columns, gpu2, 'sglang')
            summary_row[f'{gpu2} TRT Docker'] = extract_docker_from_columns(comparison_df.columns, gpu2, 'trt')

            comparison_cols = ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]
            has_valid_comparison = False
            for comp_col in comparison_cols:
                if comp_col in comparison_df.columns:
                    geomean = calculate_geomean_speedup(comparison_df, comp_col)
                    if not np.isnan(geomean):
                        summary_row[comp_col] = geomean
                        has_valid_comparison = True
                    else:
                        summary_row[comp_col] = np.nan

            # Only add to summary if we have comparative data
            if has_valid_comparison:
                summary_data.append(summary_row)
            elif verbose:
                print(f"    ⚠ Skipping summary - no comparative data between {gpu1} and {gpu2}")

            if verbose:
                print(f"    Configurations: {len(comparison_df)}")

            all_results[sheet_key] = comparison_df

    summary_df = pd.DataFrame(summary_data)

    if not summary_df.empty:
        base_cols = ['Model', 'Precision', 'Parallelism', f'{gpu1} Best', f'{gpu2} Best']
        comparison_cols = ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]
        docker_cols = [
            f'{gpu1} vLLM Docker', f'{gpu1} SGLang Docker', f'{gpu1} ATOM Docker',
            f'{gpu2} vLLM Docker', f'{gpu2} SGLang Docker', f'{gpu2} TRT Docker'
        ]

        available_cols = [col for col in base_cols + comparison_cols + docker_cols if col in summary_df.columns]
        summary_df = summary_df[available_cols]

        sort_cols = [col for col in ['Model', 'Parallelism', 'Precision'] if col in summary_df.columns]
        if sort_cols:
            summary_df = summary_df.sort_values(by=sort_cols).reset_index(drop=True)

    # Create DataFrames for version changes
    version_changes_df = pd.DataFrame(all_version_changes)
    if not version_changes_df.empty:
        sort_cols = [col for col in ['Model', 'GPU', 'Framework', 'Precision', 'Parallelism'] 
                     if col in version_changes_df.columns]
        if sort_cols:
            version_changes_df = version_changes_df.sort_values(by=sort_cols).reset_index(drop=True)

    competitive_changes_df = pd.DataFrame(all_competitive_changes)
    if not competitive_changes_df.empty:
        sort_cols = [col for col in ['Model', 'Framework', 'Precision', 'Parallelism'] 
                     if col in competitive_changes_df.columns]
        if sort_cols:
            competitive_changes_df = competitive_changes_df.sort_values(by=sort_cols).reset_index(drop=True)

    if all_results:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                ws_summary = writer.sheets['Summary']
                format_summary_sheet(ws_summary, summary_df, gpu1, gpu2)

            # Add Version Changes sheet
            if not version_changes_df.empty:
                version_changes_df.to_excel(writer, sheet_name='Version Changes', index=False)
                ws_versions = writer.sheets['Version Changes']
                format_version_changes_sheet(ws_versions, version_changes_df)
                if verbose:
                    print(f"\n  Version changes tracked: {len(version_changes_df)}")

            # Add Competitive Changes sheet
            if not competitive_changes_df.empty:
                competitive_changes_df.to_excel(writer, sheet_name='Competitive Changes', index=False)
                ws_competitive = writer.sheets['Competitive Changes']
                format_competitive_changes_sheet(ws_competitive, competitive_changes_df, gpu1, gpu2)
                if verbose:
                    print(f"  Competitive ratio changes tracked: {len(competitive_changes_df)}")

            # Add Recent Additions sheet
            if all_additions:
                # Sort by timestamp descending (most recent first)
                additions_sorted = sorted(all_additions, key=lambda x: x['timestamp'], reverse=True)
                
                # Use today's date for highlighting (not the most recent in data)
                latest_date = today_str
                
                # Convert timestamps to human-readable format for display
                def format_timestamp_display(ts):
                    """Convert '20260205-060528' to '2026-02-05 06:05:28'"""
                    if not ts:
                        return 'Unknown'
                    
                    # Check if it's a valid timestamp format (YYYYMMDD-HHMMSS)
                    match = re.match(r'^(\d{4})(\d{2})(\d{2})-(\d{2})(\d{2})(\d{2})$', ts)
                    if match:
                        year, month, day, hour, minute, second = match.groups()
                        return f"{year}-{month}-{day} {hour}:{minute}:{second}"
                    
                    # Check for merged files indicator
                    if 'merged' in ts.lower() or 'sweep' in ts.lower():
                        return 'Merged'
                    
                    # Return original if not parseable
                    return ts if len(ts) < 20 else ts[:20]
                
                # Format timestamps in the data
                for item in additions_sorted:
                    item['timestamp_display'] = format_timestamp_display(item['timestamp'])
                
                # Create DataFrame
                additions_df = pd.DataFrame(additions_sorted)
                
                # Reorder columns (use timestamp_display for human-readable format)
                # Note: version column removed as docker_image already contains version info
                col_order = ['timestamp_display', 'model', 'gpu', 'framework', 'precision', 
                            'parallelism', 'docker_image', 'num_configs', 'filename']
                additions_df = additions_df[[c for c in col_order if c in additions_df.columns]]
                
                # Rename columns for display
                additions_df = additions_df.rename(columns={
                    'timestamp_display': 'Date/Time',
                    'model': 'Model',
                    'gpu': 'GPU',
                    'framework': 'Framework',
                    'precision': 'Precision',
                    'parallelism': 'Parallelism',
                    'docker_image': 'Docker Image',
                    'num_configs': 'Configs',
                    'filename': 'File'
                })
                
                additions_df.to_excel(writer, sheet_name='All Benchmarks', index=False)
                ws_additions = writer.sheets['All Benchmarks']
                
                # Format the sheet - Header formatting
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for col_idx in range(1, len(additions_df.columns) + 1):
                    cell = ws_additions.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Highlight the most recent day's additions (the latest date in the data)
                recent_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                for row_idx, row_data in enumerate(additions_sorted, start=2):
                    if row_data['timestamp'].startswith(latest_date):
                        for col_idx in range(1, len(additions_df.columns) + 1):
                            ws_additions.cell(row=row_idx, column=col_idx).fill = recent_fill
                
                # Auto-width columns
                for col_idx, col_name in enumerate(additions_df.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = max(len(str(col_name)), 
                                 additions_df[col_name].astype(str).str.len().max() if len(additions_df) > 0 else 0)
                    ws_additions.column_dimensions[col_letter].width = min(max_len + 2, 50)
                
                ws_additions.freeze_panes = 'A2'
                
                # Count today's additions
                today_count = sum(1 for a in all_additions if a['timestamp'].startswith(today_str))
                if verbose:
                    print(f"  Total benchmarks: {len(all_additions)}, added today: {today_count}")

            used_sheet_names = {'Summary', 'Version Changes', 'Competitive Changes', 'All Benchmarks'}

            for sheet_key, df in all_results.items():
                parts = sheet_key.rsplit('_', 2)
                if len(parts) >= 3:
                    model_part = parts[0]
                    parallelism_part = parts[1]
                    precision_part = parts[2]
                else:
                    model_part = sheet_key
                    parallelism_part = 'unknown'
                    precision_part = 'unknown'

                sheet_name = _create_short_sheet_name(
                    model_part, precision_part, parallelism_part, used_sheet_names
                )
                used_sheet_names.add(sheet_name)

                df_with_summary = add_geomean_summary(df, gpu1, gpu2)
                df_with_summary.to_excel(writer, sheet_name=sheet_name, index=False)
                format_sheet_with_merged_headers(writer.sheets[sheet_name], df_with_summary, gpu1, gpu2)

        if verbose:
            print("\n" + "=" * 80)
            print(f"✓ Exported comparative analysis to: {output_file}")
            print(f"  Summary sheet with {len(summary_df)} model configurations")
            print(f"  Individual sheets: {len(all_results)}")

    if all_warnings:
        print("\n" + "=" * 80)
        print("WARNINGS:")
        for warning in all_warnings:
            print(f"  ⚠ {warning}")

    return all_results


# =============================================================================
# Main Entry Point
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Benchmark Analysis Tool - Fetch and analyze MLflow benchmark results',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process specific models
  python benchmark_analysis.py --models "Llama-3.3-70B" "DeepSeek-R1" "Qwen3-235B"

  # Process all models
  python benchmark_analysis.py --models "*"

  # With custom settings
  python benchmark_analysis.py --models "Llama*" --run-pattern "oob_" --output-dir ./results

  # Skip data fetching, only generate comparison from existing CSVs
  python benchmark_analysis.py --skip-fetch --output-dir ./existing_results

Environment Variables:
  DATABRICKS_WORKSPACE_URL  Azure Databricks workspace URL
  DATABRICKS_API_KEY        Databricks personal access token
        """
    )

    parser.add_argument(
        '--models', '-m',
        nargs='+',
        required=True,
        help='Model names to process (supports wildcards). Use "*" for all models.'
    )

    parser.add_argument(
        '--output-dir', '-o',
        default='./benchmark_results',
        help='Output directory for CSV files and Excel report (default: ./benchmark_results)'
    )

    parser.add_argument(
        '--output-excel', '-e',
        default=None,
        help='Output Excel file path (default: {output-dir}/OOB_comparative_analysis_YYYYMMDD.xlsx)'
    )

    parser.add_argument(
        '--run-pattern', '-r',
        default=None,
        help='Pattern to match parent run names. If not specified, the default run discovery behavior is used.'
    )

    parser.add_argument(
        '--gpu1',
        default='MI355X',
        help='First GPU for comparison (default: MI355X)'
    )

    parser.add_argument(
        '--gpu2',
        default='B200',
        help='Second GPU for comparison (default: B200)'
    )

    parser.add_argument(
        '--workspace-url',
        default=None,
        help='Databricks workspace URL (overrides DATABRICKS_WORKSPACE_URL env var)'
    )

    parser.add_argument(
        '--api-key',
        default=None,
        help='Databricks API key (overrides DATABRICKS_API_KEY env var)'
    )

    parser.add_argument(
        '--tag', '-t',
        default=None,
        help='Filter runs by tag (e.g., "aws_perf=true"). Can be used instead of or with --run-pattern'
    )

    parser.add_argument(
        '--hostname',
        default=None,
        help='Filter runs by server hostname (matches server/misc.hostname parameter)'
    )

    parser.add_argument(
        '--skip-fetch',
        action='store_true',
        help='Skip fetching data from Databricks, only generate comparison from existing CSVs'
    )

    parser.add_argument(
        '--no-merge',
        action='store_true',
        help='Do not merge sweeps with same configuration'
    )

    parser.add_argument(
        '--quiet', '-q',
        action='store_true',
        help='Suppress verbose output'
    )

    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Generate output filename with date: OOB_comparative_analysis_YYYYMMDD.xlsx
    date_str = datetime.now().strftime('%Y%m%d')
    default_excel_name = f'OOB_comparative_analysis_{date_str}.xlsx'
    output_excel = args.output_excel or str(output_dir / default_excel_name)

    verbose = not args.quiet

    if verbose:
        print("=" * 80)
        print("BENCHMARK ANALYSIS TOOL")
        print("=" * 80)
        print(f"Models: {args.models}")
        print(f"Output directory: {output_dir}")
        print(f"Output Excel: {output_excel}")
        if args.run_pattern:
            print(f"Run pattern: {args.run_pattern}")
        if args.tag:
            print(f"Tag filter: {args.tag}")
        if args.hostname:
            print(f"Hostname filter: {args.hostname}")
        print(f"GPUs: {args.gpu1} vs {args.gpu2}")
        print("=" * 80)

    # Step 1: Fetch data from Databricks (unless skipped)
    if not args.skip_fetch:
        workspace_url = args.workspace_url or os.environ.get('DATABRICKS_WORKSPACE_URL')
        api_key = args.api_key or os.environ.get('DATABRICKS_API_KEY')

        if not workspace_url or not api_key:
            print("Error: Databricks credentials required.")
            print("Set DATABRICKS_WORKSPACE_URL and DATABRICKS_API_KEY environment variables,")
            print("or use --workspace-url and --api-key arguments.")
            return 1

        if verbose:
            print("\nConnecting to Databricks...")

        client = DatabricksMLflowClient(workspace_url, api_key)

        for model_pattern in args.models:
            if model_pattern == '*':
                experiment_pattern = '*'
            else:
                experiment_pattern = f'*{model_pattern}*'

            if verbose:
                print(f"\nProcessing models matching: {experiment_pattern}")

            summary = process_all_experiments(
                client=client,
                parent_run_pattern=args.run_pattern,
                experiment_name_pattern=experiment_pattern,
                output_dir=str(output_dir),
                merge_same_config=not args.no_merge,
                tag_filter=args.tag,
                hostname_filter=args.hostname,
                verbose=verbose
            )

            if verbose:
                print(f"\nProcessed {len(summary['processed'])} experiments")
                if summary['failed']:
                    print(f"Failed: {len(summary['failed'])}")

    # Step 2: Generate comparative analysis Excel
    if verbose:
        print("\n" + "=" * 80)
        print("GENERATING COMPARATIVE ANALYSIS")
        print("=" * 80)

    try:
        results = create_comparative_analysis_enhanced(
            parent_dir=str(output_dir),
            output_file=output_excel,
            gpu1=args.gpu1,
            gpu2=args.gpu2,
            verbose=verbose
        )

        if verbose and results:
            print(f"\n✓ Analysis complete! Output: {output_excel}")

    except Exception as e:
        print(f"Error generating comparative analysis: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == '__main__':
    exit(main())
