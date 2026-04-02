# import glob
import os
import re
import requests
# from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from IPython.display import display


class DatabricksMLflowClient:
    """Client for interacting with Azure Databricks MLflow API"""

    def __init__(self, workspace_url: str, api_key: str):
        """
        Initialize Databricks MLflow client

        Args:
            workspace_url: Azure Databricks workspace URL (e.g., 'https://adb-xxxxx.azuredatabricks.net')
            api_key: Databricks personal access token
        """
        self.workspace_url = workspace_url.rstrip('/')
        self.api_key = api_key
        self.headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json'
        }
        self.mlflow_base_url = f"{self.workspace_url}/api/2.0/mlflow"

    def list_experiments(self, max_results: int = 1000) -> List[Dict]:
        """List all experiments"""
        url = f"{self.mlflow_base_url}/experiments/search"

        all_experiments = []
        page_token = None

        while True:
            params = {'max_results': max_results}
            if page_token:
                params['page_token'] = page_token

            response = requests.get(url, headers=self.headers, params=params)
            response.raise_for_status()

            data = response.json()
            experiments = data.get('experiments', [])
            all_experiments.extend(experiments)

            page_token = data.get('next_page_token')
            if not page_token:
                break

        return all_experiments

    def get_experiment(self, experiment_id: str) -> Dict:
        """Get experiment by ID"""
        url = f"{self.mlflow_base_url}/experiments/get"
        params = {'experiment_id': experiment_id}

        response = requests.get(url, headers=self.headers, params=params)
        response.raise_for_status()

        return response.json()['experiment']

    def get_experiment_by_name(self, experiment_name: str) -> Optional[Dict]:
        """
        Get experiment by name

        Args:
            experiment_name: Name of the experiment to find

        Returns:
            Experiment dictionary if found, None otherwise
        """
        experiments = self.list_experiments()
        for exp in experiments:
            if exp['name'] == experiment_name:
                return exp
        return None

    def search_runs(
        self,
        experiment_ids: Optional[List[str]] = None,
        filter_string: Optional[str] = None,
        run_name: Optional[str] = None,
        max_results: int = 1000,
        order_by: Optional[List[str]] = None
    ) -> List[Dict]:
        """
        Search runs across experiments

        Args:
            experiment_ids: List of experiment IDs to search (if None, searches all experiments)
            filter_string: MLflow search filter (e.g., "metrics.accuracy > 0.9")
            run_name: Filter by run name (exact match or partial with wildcards)
            max_results: Maximum results per page
            order_by: List of order by clauses (e.g., ["metrics.accuracy DESC"])
        """
        url = f"{self.mlflow_base_url}/runs/search"

        # Build filter string with run name if provided
        if run_name:
            name_filter = f"tags.mlflow.runName = '{run_name}'"
            if filter_string:
                filter_string = f"{filter_string} AND {name_filter}"
            else:
                filter_string = name_filter

        all_runs = []
        page_token = None

        while True:
            payload = {
                'max_results': max_results
            }

            # Only add experiment_ids if provided
            if experiment_ids:
                payload['experiment_ids'] = experiment_ids

            if filter_string:
                payload['filter'] = filter_string
            if order_by:
                payload['order_by'] = order_by
            if page_token:
                payload['page_token'] = page_token

            response = requests.post(url, headers=self.headers, json=payload)
            response.raise_for_status()

            data = response.json()
            runs = data.get('runs', [])
            all_runs.extend(runs)

            page_token = data.get('next_page_token')
            if not page_token:
                break

        return all_runs

    def get_run(self, run_id: str) -> Dict:
        """Get run details by ID"""
        url = f"{self.mlflow_base_url}/runs/get"
        params = {'run_id': run_id}

        response = requests.get(url, headers=self.headers, params=params)
        response.raise_for_status()

        return response.json()['run']

    @staticmethod
    def _extract_dict_or_list(data: Dict, key: str) -> Dict:
        """
        Extract dictionary from run data, handling both dict and list formats

        Args:
            data: Run data dictionary from MLflow API
            key: Key to extract ('params', 'metrics', or 'tags')

        Returns:
            Dictionary with key-value pairs
        """
        items = data.get(key, {})

        # If items is a list, convert to dict
        if isinstance(items, list):
            items_dict = {}
            for item in items:
                if 'key' in item and 'value' in item:
                    items_dict[item['key']] = item['value']
            return items_dict

        return items

    @staticmethod
    def _extract_tags(run_data: Dict) -> Dict:
        """
        Extract tags from run data, handling both dict and list formats

        Args:
            run_data: Run data dictionary from MLflow API

        Returns:
            Tags as dictionary
        """
        return DatabricksMLflowClient._extract_dict_or_list(run_data, 'tags')

    def list_runs_by_name(
        self,
        experiment_ids: Optional[List[str]] = None,
        run_name_pattern: Optional[str] = None,
        exact_match: bool = False,
        max_results: int = 1000
    ) -> List[Dict]:
        """
        List runs filtered by name pattern

        Args:
            experiment_ids: List of experiment IDs to search (if None, searches all)
            run_name_pattern: Pattern to match run names against
            exact_match: If True, exact match; if False, contains match
            max_results: Maximum results to return

        Returns:
            List of matching runs
        """
        if run_name_pattern and exact_match:
            return self.search_runs(
                experiment_ids=experiment_ids,
                run_name=run_name_pattern,
                max_results=max_results
            )

        all_runs = self.search_runs(
            experiment_ids=experiment_ids,
            max_results=max_results
        )

        if run_name_pattern:
            filtered_runs = []
            for run in all_runs:
                tags = self._extract_tags(run['data'])
                run_name = tags.get('mlflow.runName', '')
                if run_name_pattern in run_name:
                    filtered_runs.append(run)
            return filtered_runs

        return all_runs

    def get_run_by_name(
        self,
        run_name: str,
        experiment_ids: Optional[List[str]] = None
    ) -> Optional[Dict]:
        """
        Get a specific run by exact name match

        Args:
            run_name: Exact name of the run to find
            experiment_ids: List of experiment IDs to search (if None, searches all)

        Returns:
            Run dictionary if found, None otherwise
        """
        runs = self.list_runs_by_name(
            experiment_ids=experiment_ids,
            run_name_pattern=run_name,
            exact_match=True
        )

        if runs:
            if len(runs) > 1:
                print(f"Warning: Found {len(runs)} runs with name '{run_name}'. Returning first one.")
            return runs[0]

        return None

    def get_child_runs(
        self,
        parent_run_id: str,
        max_results: int = 1000
    ) -> List[Dict]:
        """
        Get all child runs for a parent run

        Args:
            parent_run_id: ID of the parent run
            max_results: Maximum results to return

        Returns:
            List of child run dictionaries
        """
        # Get the parent run to find its experiment
        parent_run = self.get_run(parent_run_id)
        experiment_id = parent_run['info']['experiment_id']

        filter_string = f"tags.mlflow.parentRunId = '{parent_run_id}'"
        return self.search_runs(
            experiment_ids=[experiment_id],
            filter_string=filter_string,
            max_results=max_results
        )

    def get_run_hierarchy(self, parent_run_id: str) -> Dict:
        """
        Get parent run with all its children

        Args:
            parent_run_id: ID of the parent run

        Returns:
            Dictionary with parent and children runs
        """
        parent_run = self.get_run(parent_run_id)
        child_runs = self.get_child_runs(parent_run_id)

        return {
            'parent': parent_run,
            'children': child_runs
        }

    def list_parent_runs(
        self,
        experiment_ids: Optional[List[str]] = None,
        run_name_pattern: Optional[str] = None
    ) -> List[Dict]:
        """
        List only parent runs (runs without a parent)

        Args:
            experiment_ids: List of experiment IDs to search
            run_name_pattern: Pattern to match run names

        Returns:
            List of parent runs
        """
        all_runs = self.list_runs_by_name(
            experiment_ids=experiment_ids,
            run_name_pattern=run_name_pattern,
            exact_match=False
        )

        # Filter for parent runs (those without mlflow.parentRunId tag)
        parent_runs = []
        for run in all_runs:
            tags = self._extract_tags(run['data'])
            if 'mlflow.parentRunId' not in tags:
                parent_runs.append(run)

        return parent_runs

    def get_sweep_results(
        self,
        parent_run_name: str,
        experiment_ids: Optional[List[str]] = None
    ) -> pd.DataFrame:
        """
        Get all results from a sweep (parent run and all children)

        Args:
            parent_run_name: Name of the parent/sweep run
            experiment_ids: List of experiment IDs to search

        Returns:
            DataFrame with all sweep results
        """
        # Find parent run by name
        parent_run = self.get_run_by_name(
            run_name=parent_run_name,
            experiment_ids=experiment_ids
        )

        if not parent_run:
            raise ValueError(f"Parent run '{parent_run_name}' not found")

        parent_run_id = parent_run['info']['run_id']

        # Get all child runs
        child_runs = self.get_child_runs(parent_run_id)

        # Convert to DataFrame
        df = self.runs_to_dataframe(child_runs)

        # Add parent run info
        parent_tags = self._extract_tags(parent_run['data'])
        df['parent_run_id'] = parent_run_id
        df['parent_run_name'] = parent_tags.get('mlflow.runName', 'N/A')

        return df

    def search_runs_in_experiment(
        self,
        experiment_name: str,
        filter_string: Optional[str] = None,
        max_results: int = 1000,
        order_by: Optional[List[str]] = None
    ) -> List[Dict]:
        """
        Search runs in a specific experiment by name

        Args:
            experiment_name: Name of the experiment
            filter_string: MLflow search filter
            max_results: Maximum results to return
            order_by: List of order by clauses

        Returns:
            List of runs in the experiment
        """
        experiment = self.get_experiment_by_name(experiment_name)

        if not experiment:
            raise ValueError(f"No experiment found with name: {experiment_name}")

        return self.search_runs(
            experiment_ids=[experiment['experiment_id']],
            filter_string=filter_string,
            max_results=max_results,
            order_by=order_by
        )

    @staticmethod
    def display_run_info(run: Dict):
        """
        Display formatted run information

        Args:
            run: Run dictionary from MLflow API
        """
        info = run['info']
        data = run['data']

        # Extract data properly
        tags = DatabricksMLflowClient._extract_dict_or_list(data, 'tags')
        params = DatabricksMLflowClient._extract_dict_or_list(data, 'params')
        metrics = DatabricksMLflowClient._extract_dict_or_list(data, 'metrics')

        print(f"Run ID: {info['run_id']}")
        print(f"Run Name: {tags.get('mlflow.runName', 'N/A')}")
        print(f"Status: {info['status']}")
        print(f"Start Time: {pd.to_datetime(info['start_time'], unit='ms')}")
        print(f"End Time: {pd.to_datetime(info['end_time'], unit='ms') if info.get('end_time') else 'Running'}")

        print("\nParameters:")
        for key, value in params.items():
            print(f"  {key}: {value}")

        print("\nMetrics:")
        for key, value in metrics.items():
            print(f"  {key}: {value}")

        print("\nTags:")
        for key, value in tags.items():
            if not key.startswith('mlflow.'):
                print(f"  {key}: {value}")

    @staticmethod
    def runs_to_dataframe(runs: List[Dict]) -> pd.DataFrame:
        """
        Convert runs to DataFrame for easy analysis

        Args:
            runs: List of run dictionaries from MLflow API

        Returns:
            DataFrame with run information
        """
        data = []

        for run in runs:
            info = run['info']
            run_data = run['data']

            # Extract tags, params, and metrics properly
            tags = DatabricksMLflowClient._extract_dict_or_list(run_data, 'tags')
            params = DatabricksMLflowClient._extract_dict_or_list(run_data, 'params')
            metrics = DatabricksMLflowClient._extract_dict_or_list(run_data, 'metrics')

            row = {
                'run_id': info['run_id'],
                'run_name': tags.get('mlflow.runName', 'N/A'),
                'status': info['status'],
                'start_time': pd.to_datetime(info['start_time'], unit='ms'),
                'experiment_id': info['experiment_id']
            }

            # Add end time if exists
            if info.get('end_time'):
                row['end_time'] = pd.to_datetime(info['end_time'], unit='ms')

            # Add parameters
            for key, value in params.items():
                row[f'param_{key}'] = value

            # Add metrics
            for key, value in metrics.items():
                row[f'metric_{key}'] = value

            # Add custom tags (exclude mlflow. tags)
            for key, value in tags.items():
                if not key.startswith('mlflow.'):
                    row[f'tag_{key}'] = value

            data.append(row)

        return pd.DataFrame(data)

    @staticmethod
    def sweep_results_summary(df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict]:
        """
        Create summary statistics for sweep results

        Args:
            df: DataFrame from get_sweep_results

        Returns:
            Tuple of (summary statistics DataFrame, best values dict)
        """
        # Get metric columns
        metric_cols = [col for col in df.columns if col.startswith('metric_')]

        if not metric_cols:
            return pd.DataFrame({'message': ['No metrics found']}), {}

        summary = df[metric_cols].describe()

        # Add best values
        best_values = {}
        for col in metric_cols:
            best_idx = df[col].idxmax()
            best_values[col] = {
                'best_value': df.loc[best_idx, col],
                'best_run': df.loc[best_idx, 'run_name'],
                'best_run_id': df.loc[best_idx, 'run_id']
            }

        return summary, best_values


def postprocess_sweep_results(
    df: pd.DataFrame,
    select_columns: Optional[List[str]] = None
) -> pd.DataFrame:
    """
    Postprocess sweep results into a clean table

    Args:
        df: DataFrame from get_sweep_results
        select_columns: List of columns to keep. If None, uses default set.

    Returns:
        Cleaned and sorted DataFrame
    """
    # Create a copy to avoid modifying original
    df_clean = df.copy()

    # Rename columns by stripping the prefix before '/'
    rename_map = {}
    seen_names = set()

    for col in df_clean.columns:
        if col.startswith('param_') or col.startswith('metric_'):
            # Extract the part after the first '/'
            if '/' in col:
                new_name = col.split('/', 1)[1]

                # Handle duplicates by keeping only the first occurrence
                if new_name not in seen_names:
                    rename_map[col] = new_name
                    seen_names.add(new_name)

    df_clean = df_clean.rename(columns=rename_map)

    # Remove any remaining duplicate columns (keep first)
    df_clean = df_clean.loc[:, ~df_clean.columns.duplicated()]

    # Compute median_e2el_ms if not present
    if 'median_e2el_ms' not in df_clean.columns:
        required_cols = ['median_ttft_ms', 'median_itl_ms', 'total_output_tokens']
        if all(col in df_clean.columns for col in required_cols):
            df_clean['median_e2el_ms'] = (
                df_clean['median_ttft_ms'] +
                df_clean['median_itl_ms'] * df_clean['total_output_tokens'].astype(float)
            )

    # If median_e2el_ms exists but has NaN values, fill them
    if 'median_e2el_ms' in df_clean.columns:
        required_cols = ['median_ttft_ms', 'median_itl_ms', 'total_output_tokens']
        if all(col in df_clean.columns for col in required_cols):
            df_clean['median_e2el_ms'] = df_clean['median_e2el_ms'].fillna(
                df_clean['median_ttft_ms'] +
                df_clean['median_itl_ms'] * df_clean['total_output_tokens'].astype(float)
            )

    # Identify and convert sort columns to numeric
    sort_columns = []
    for col_name in ['random_input_len', 'random_output_len', 'max_concurrency']:
        if col_name in df_clean.columns:
            try:
                # Convert to numeric
                df_clean[col_name] = pd.to_numeric(df_clean[col_name], errors='coerce')
                sort_columns.append(col_name)
            except Exception as e:
                print(f"Warning: Could not convert column '{col_name}' to numeric: {e}")

    # Sort by the columns in order: random_input_len, random_output_len, max_concurrency
    if sort_columns:
        try:
            df_clean = df_clean.sort_values(by=sort_columns, ascending=True)
        except Exception as e:
            print(f"Warning: Could not sort by {sort_columns}: {e}")

    # Reset index after sorting
    df_clean = df_clean.reset_index(drop=True)

    # Select default columns if not specified
    if select_columns is None:
        default_columns = [
            'experiment_url',
            'backend',
            'model',
            'random_input_len',
            'random_output_len',
            'max_concurrency',
            'median_itl_ms',
            'median_ttft_ms',
            'median_tpot_ms',
            'median_e2el_ms',
            'output_throughput',
            'total_token_throughput',
            'status'
        ]
        # Filter to columns that actually exist
        select_columns = [col for col in default_columns if col in df_clean.columns]

    # Select the specified columns
    if select_columns:
        # Keep only columns that exist
        available_cols = [col for col in select_columns if col in df_clean.columns]
        if available_cols:
            df_clean = df_clean[available_cols]

    return df_clean

def create_benchmark_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create a focused summary table with key benchmark metrics

    Args:
        df: DataFrame from postprocess_sweep_results

    Returns:
        Summary DataFrame with key columns
    """
    # Key columns to include
    key_columns = [
        'run_name',
        'status',
        'backend',
        'model',
        'random_input_len',
        'random_output_len',
        'max_concurrency',
        'mean_itl_ms',
        'mean_ttft_ms',
        'mean_tpot_ms'
    ]

    # Filter to columns that exist
    available_columns = [col for col in key_columns if col in df.columns]

    # Also include any other metric columns
    metric_columns = [col for col in df.columns if col not in available_columns
                     and not col.startswith('run_')
                     and not col.startswith('experiment_')
                     and not col.startswith('parent_')
                     and not col.startswith('start_')
                     and not col.startswith('end_')
                     and col not in ['backend', 'model']]

    # Combine columns
    summary_columns = available_columns + metric_columns

    # Create summary
    df_summary = df[summary_columns].copy()

    # Rename for clarity
    rename_map = {
        'random_input_len': 'input_len',
        'random_output_len': 'output_len',
        'max_concurrency': 'concurrency'
    }

    df_summary = df_summary.rename(columns={k: v for k, v in rename_map.items() if k in df_summary.columns})

    return df_summary


def compare_configurations(df: pd.DataFrame,
                          metric: str = 'mean_tpot_ms',
                          group_by: Optional[List[str]] = None) -> pd.DataFrame:
    """
    Compare performance across different configurations

    Args:
        df: Postprocessed DataFrame
        metric: Metric to compare (default: mean_tpot_ms)
        group_by: Columns to group by (default: ['random_input_len', 'random_output_len'])

    Returns:
        Pivot table comparing configurations
    """
    if group_by is None:
        group_by = ['random_input_len', 'random_output_len']

    # Filter to columns that exist
    group_by = [col for col in group_by if col in df.columns]

    if not group_by or metric not in df.columns:
        print(f"Warning: Required columns not found")
        return pd.DataFrame()

    # Create pivot table
    if 'max_concurrency' in df.columns:
        pivot = df.pivot_table(
            values=metric,
            index=group_by,
            columns='max_concurrency',
            aggfunc='mean'
        )
        pivot.columns.name = 'Concurrency'
        pivot = pivot.round(2)
        return pivot
    else:
        # Just group and aggregate
        return df.groupby(group_by)[metric].agg(['mean', 'min', 'max', 'std']).round(2)


def export_sweep_results(
    df: pd.DataFrame,
    filename: str,
    metadata: Optional[Dict[str, str]] = None,
    filter_finished: bool = True,
    drop_status: bool = True,
    verbose: bool = True
) -> None:
    """
    Export sweep results to CSV, optionally filtering for finished runs only

    Args:
        df: DataFrame from postprocess_sweep_results
        filename: Output CSV filename (e.g., 'results.csv')
        metadata: Dictionary of metadata to add as columns (e.g., docker_image, gpu, framework)
        filter_finished: If True, only export runs with status='FINISHED'
        drop_status: If True, remove status column from export
        verbose: If True, print export statistics
    """
    df_export = df.copy()

    # Track statistics
    total_runs = len(df_export)

    # Filter for finished runs if requested
    if filter_finished and 'status' in df_export.columns:
        finished_mask = df_export['status'] == 'FINISHED'
        failed_count = (~finished_mask).sum()

        if verbose:
            print(f"Total runs: {total_runs}")
            print(f"Finished runs: {finished_mask.sum()}")
            print(f"Failed/Other runs: {failed_count}")

        if failed_count > 0:
            # Show details about non-finished runs
            non_finished_df = df_export[~finished_mask]
            status_counts = non_finished_df['status'].value_counts()

            print(f"\nFiltering out {failed_count} non-FINISHED run(s):")
            for status, count in status_counts.items():
                print(f"  - {status}: {count}")

            # Show detailed info for each non-finished run
            config_cols = ['random_input_len', 'random_output_len', 'max_concurrency']
            available_config_cols = [col for col in config_cols if col in non_finished_df.columns]

            print(f"  Runs affected:")
            for _, row in non_finished_df.iterrows():
                run_status = row['status']

                # Build config string (may be empty if run failed early)
                config_parts = []
                for col in available_config_cols:
                    if pd.notna(row.get(col)):
                        config_parts.append(f"{col}={row[col]}")
                config_str = ", ".join(config_parts) if config_parts else "(no config - run failed before parameters were set)"

                # Try to get experiment URL for easy debugging
                exp_url = row.get('experiment_url', '')
                if exp_url and pd.notna(exp_url):
                    print(f"    - [{run_status}] {config_str}")
                    print(f"      URL: {exp_url}")
                else:
                    print(f"    - [{run_status}] {config_str}")

        df_export = df_export[finished_mask].copy()

        if len(df_export) == 0:
            print("Warning: No FINISHED runs to export!")
            return

    # Add metadata columns at the beginning if provided
    if metadata:
        # Insert metadata columns at the start (reversed order to maintain order)
        for key in reversed(list(metadata.keys())):
            df_export.insert(0, key, metadata[key])

    # Drop status column if requested
    if drop_status and 'status' in df_export.columns:
        df_export = df_export.drop(columns=['status'])
        if verbose:
            print(f"Removed 'status' column from export")

    # Reset index
    df_export = df_export.reset_index(drop=True)

    # Export to CSV
    df_export.to_csv(filename, index=False)

    if verbose:
        print(f"\n✓ Exported {len(df_export)} runs to: {filename}")
        print(f"Columns: {list(df_export.columns)}")


def parse_gpu_model(gpu_string: str, docker_runtime: str = None, docker_image: str = None) -> str:
    """
    Parse GPU model from server logs

    Args:
        gpu_string: GPU identifier (e.g., "NVIDIA H100", "(gfx950:sramecc+:xnack-)")
        docker_runtime: Docker runtime (e.g., "nvidia", "rocm")
        docker_image: Docker image name (e.g., "rocm/vllm:...")

    Returns:
        Cleaned GPU model name
    """
    if not gpu_string or pd.isna(gpu_string):
        # Default based on docker runtime
        if docker_runtime and 'nvidia' in str(docker_runtime).lower():
            return 'B200'
        if docker_runtime and 'rocm' in str(docker_runtime).lower():
            return 'MI355X'

        # Default based on docker image name
        if docker_image:
            docker_str = str(docker_image).lower()
            if 'rocm' in docker_str:
                return 'MI355X'
            if 'nvidia' in docker_str or 'nvcr' in docker_str:
                return 'B200'

        return 'unknown'

    gpu_str = str(gpu_string).strip()

    # Handle AMD GPUs like "(gfx950:sramecc+:xnack-)"
    if 'gfx950' in gpu_str:
        return 'MI355X'
    elif 'gfx942' in gpu_str:
        return 'MI300X'
    elif 'gfx90a' in gpu_str:
        return 'MI250X'
    elif 'gfx' in gpu_str:
        # Extract gfx number for other AMD GPUs
        gfx_match = re.search(r'gfx(\d+)', gpu_str)
        if gfx_match:
            return f'AMD_GFX{gfx_match.group(1)}'

    # Handle NVIDIA GPUs
    if 'H100' in gpu_str:
        return 'H100'
    elif 'H200' in gpu_str:
        return 'H200'
    elif 'B200' in gpu_str:
        return 'B200'
    elif 'A100' in gpu_str:
        return 'A100'
    elif 'V100' in gpu_str:
        return 'V100'

    # Generic cleanup
    gpu_str = gpu_str.replace('NVIDIA', '').replace('AMD', '')
    gpu_str = re.sub(r'[^\w\s-]', '', gpu_str).strip()

    return gpu_str if gpu_str else 'unknown'

def extract_framework_from_docker_image(docker_image: str) -> str:
    """
    Extract framework from docker image name

    Args:
        docker_image: Docker image (e.g., "rocm/vllm:rocm7.0.0_vllm_0.11.2_20251210")

    Returns:
        Framework name (vllm, sglang, atom, trt, or unknown)
    """
    if not docker_image or pd.isna(docker_image):
        return 'unknown'

    docker_str = str(docker_image).lower()

    # Check for TensorRT-LLM (various naming patterns)
    if 'tensorrt-llm' in docker_str or 'tensorrt_llm' in docker_str or 'trtllm' in docker_str or 'tritonserver' in docker_str:
        return 'trt'

    # Check for SGLang
    if 'sglang' in docker_str:
        return 'sglang'

    # Check for Atom
    if 'atom' in docker_str:
        return 'atom'

    # Check for vLLM (should be last since it's most common)
    if 'vllm' in docker_str:
        return 'vllm'

    return 'unknown'


def extract_docker_image(docker_image_string: str) -> str:
    """
    Extract and clean docker image name

    Args:
        docker_image_string: Docker image (e.g., "rocm/vllm:rocm7.0.0_vllm_0.11.2_20251210")

    Returns:
        Cleaned docker image name
    """
    if not docker_image_string or pd.isna(docker_image_string):
        return 'unknown'

    return str(docker_image_string).strip()


def extract_model_name_from_params(df: pd.DataFrame) -> str:
    """
    Extract model name from job parameters

    Args:
        df: DataFrame with sweep results (before postprocessing)

    Returns:
        Model name from parameters, or 'unknown' if not found
    """
    # Try different parameter columns in order of preference
    param_cols = [
        'param_server/process.args.model-path',
        'param_vllm_bench_serve/model',
        'param_server/model',
        'param_model'
    ]

    for col in param_cols:
        if col in df.columns:
            # Get the first non-null value
            model_values = df[col].dropna()
            if len(model_values) > 0:
                model_name = str(model_values.iloc[0]).strip()
                if model_name and model_name.lower() != 'none':
                    # Strip org prefix (e.g., "zai-org/GLM-4.7-FP8" -> "GLM-4.7-FP8")
                    if '/' in model_name:
                        model_name = model_name.split('/')[-1]
                    return model_name

    return 'unknown'


def normalize_model_name(model_name: str) -> str:
    """
    Normalize model name to canonical form for grouping

    Args:
        model_name: Model name (e.g., "GLM-4.7-FP8", "nvidia-Llama-3.3-70B-Instruct-NVFP4")

    Returns:
        Normalized model name (e.g., "GLM-4.7", "Llama-3.3-70B")
    """
    # Remove any remaining slashes (shouldn't happen but be safe)
    model_str = model_name.replace('/', '-')

    # Remove vendor prefixes (nvidia-, amd-, zai-org-, etc.)
    model_str = re.sub(r'^(nvidia|amd|zai-org|huggingface)-', '', model_str, flags=re.IGNORECASE)

    # Remove precision suffixes (FP8, FP4, NVFP4, MXFP4, etc.)
    model_str = re.sub(r'-(NV|MX)?FP\d+.*$', '', model_str, flags=re.IGNORECASE)

    # Remove other common suffixes
    model_str = re.sub(r'-(Preview|Instruct|Base|Chat)$', '', model_str, flags=re.IGNORECASE)

    # Clean up any trailing dashes
    model_str = model_str.strip('-')

    return model_str

def extract_precision_from_name(name: str) -> str:
    """
    Extract precision from experiment or model name
    
    Args:
        name: Experiment or model name (e.g., "amd--Llama-3.3-70B-Instruct-FP8-KV")
    
    Returns:
        Precision string (mxfp4, fp8, fp4, bf16, etc.) or 'unknown'
    """
    name_lower = name.lower()
    
    # Check for common precision patterns
    # Note: check for the more specific "mxfp4" first so it is not
    # accidentally grouped with generic fp4 patterns.
    if 'mxfp4' in name_lower:
        return 'mxfp4'
    if 'fp4' in name_lower:
        return 'fp4'
    if 'fp8' in name_lower:
        return 'fp8'
    if 'bf16' in name_lower:
        return 'bf16'
    if 'fp16' in name_lower:
        return 'fp16'
    if 'int8' in name_lower:
        return 'int8'
    if 'int4' in name_lower:
        return 'int4'
    
    return 'unknown'


def extract_parallelism_from_params(df: pd.DataFrame) -> str:
    """
    Extract parallelism configuration from run parameters
    
    Args:
        df: DataFrame with sweep results
    
    Returns:
        Parallelism string (e.g., 'tp8', 'tp8ep4') or 'unknown'
    """
    # Try to find tensor_parallel_size in parameters
    tp_cols = [col for col in df.columns if 'tensor_parallel' in col.lower() or 'tp_size' in col.lower()]
    ep_cols = [col for col in df.columns if 'expert_parallel' in col.lower() or 'ep_size' in col.lower()]
    
    tp_size = None
    ep_size = None
    
    for col in tp_cols:
        if col in df.columns and not df[col].isna().all():
            try:
                tp_size = int(float(df[col].iloc[0]))
                break
            except (ValueError, TypeError):
                pass
    
    for col in ep_cols:
        if col in df.columns and not df[col].isna().all():
            try:
                ep_size = int(float(df[col].iloc[0]))
                break
            except (ValueError, TypeError):
                pass
    
    if tp_size and ep_size:
        return f'tp{tp_size}ep{ep_size}'
    elif tp_size:
        return f'tp{tp_size}'
    
    return 'unknown'


def parse_experiment_config(experiment_name: str, exp_name: str = None, df: pd.DataFrame = None) -> dict:
    """
    Parse experiment/run name to extract configuration

    Args:
        experiment_name: Parent run name like "oob_vllm_tp8_fp8" or "aws_perf_benchmark_sweep"
        exp_name: Optional MLflow experiment name for fallback precision extraction
        df: Optional DataFrame for extracting parallelism from parameters

    Returns:
        Dictionary with framework, parallelism, precision
    """
    # Pattern: oob_{framework}_{parallelism}_{precision}
    # Handle both tp8 and tp8ep8 patterns
    pattern = r'oob_([a-z]+)_(tp\d+(?:ep\d+)?)_([a-z0-9]+)'
    match = re.match(pattern, experiment_name)

    if match:
        return {
            'framework': match.group(1),
            'parallelism': match.group(2),
            'precision': match.group(3)
        }

    # Fallback: try simpler pattern without parallelism details
    simple_pattern = r'oob_([a-z]+)_.*_([a-z0-9]+)$'
    match = re.match(simple_pattern, experiment_name)

    if match:
        return {
            'framework': match.group(1),
            'parallelism': 'unknown',
            'precision': match.group(2)
        }

    # If pattern didn't match, try to extract from experiment name and parameters
    result = {
        'framework': 'unknown',
        'parallelism': 'unknown',
        'precision': 'unknown'
    }

    def _get_precision_from_sources(exp_name, experiment_name, df):
        """
        Determine precision from experiment and run names, falling back to DataFrame parameters.
        """
        # Try to get precision from experiment name
        if exp_name:
            precision = extract_precision_from_name(exp_name)
            if precision != 'unknown':
                return precision

        # If still unknown, try the run name itself
        precision = extract_precision_from_name(experiment_name)
        if precision != 'unknown':
            return precision

        # Try to get precision from DataFrame parameters (e.g., quantization settings)
        if df is not None and not df.empty:
            # Check for model name columns that might contain precision
            model_cols = ['param_server/process.args.model-path', 'param_server/model', 'param_model']
            for col in model_cols:
                if col in df.columns and not df[col].isna().all():
                    model_path = str(df[col].iloc[0])
                    precision = extract_precision_from_name(model_path)
                    if precision != 'unknown':
                        return precision

            # Also check for explicit quantization/dtype parameters
            quant_cols = [
                'param_quantization', 'param_dtype', 'param_kv_cache_dtype',
                'param_server/process.args.quantization',
                'param_server/process.args.dtype',
                'param_server/process.args.kv-cache-dtype',
                'param_server/vllm.quantization',
                'param_server/vllm.dtype',
            ]
            for col in quant_cols:
                if col in df.columns and not df[col].isna().all():
                    quant_val = str(df[col].iloc[0]).lower()
                    if 'fp8' in quant_val or 'float8' in quant_val:
                        return 'fp8'
                    elif 'fp4' in quant_val:
                        return 'fp4'
                    elif 'bf16' in quant_val or 'bfloat16' in quant_val:
                        return 'bf16'
                    elif 'fp16' in quant_val or 'float16' in quant_val:
                        return 'fp16'

        return 'unknown'

    # Use helper to determine precision from all available sources
    result['precision'] = _get_precision_from_sources(exp_name, experiment_name, df)
    
    # Try to get parallelism from DataFrame parameters
    if df is not None and not df.empty:
        parallelism = extract_parallelism_from_params(df)
        if parallelism != 'unknown':
            result['parallelism'] = parallelism
    
    return result


def format_timestamp(timestamp_ms: int) -> str:
    """
    Format MLflow timestamp (milliseconds since epoch) to YYYYMMDD-HHMMSS

    Args:
        timestamp_ms: Timestamp in milliseconds

    Returns:
        Formatted timestamp string
    """
    dt = datetime.fromtimestamp(timestamp_ms / 1000.0)
    return dt.strftime('%Y%m%d-%H%M%S')


def extract_failed_requests(df: pd.DataFrame) -> int:
    """
    Extract number of failed requests from sweep results

    Args:
        df: DataFrame from get_sweep_results (raw data)

    Returns:
        Total number of failed requests across all runs
    """
    failed_col = 'metric_genai-perf/failed'
    if failed_col in df.columns:
        return int(df[failed_col].sum())
    return 0

def process_all_experiments(
    client: DatabricksMLflowClient,
    parent_run_pattern: str = "oob_",
    experiment_name_pattern: Optional[str] = None,
    output_dir: str = ".",
    merge_same_config: bool = True,
    tag_filter: Optional[str] = None,
    hostname_filter: Optional[str] = None,
    verbose: bool = True
) -> dict:
    """
    Process all experiments with matching parent run names and export results

    Args:
        client: DatabricksMLflowClient instance
        parent_run_pattern: Pattern to match parent run names
        experiment_name_pattern: Optional pattern to filter experiment names
        output_dir: Directory to save CSV files
        merge_same_config: If True, merge multiple sweeps with same config into one CSV
        tag_filter: Optional tag filter (e.g., "aws_perf=true")
        hostname_filter: Optional hostname filter (matches server/misc.hostname parameter)
        verbose: Print progress information

    Returns:
        Dictionary with processing summary
    """
    import fnmatch
    from collections import defaultdict

    # Get all experiments
    all_experiments = client.list_experiments()

    # Filter experiments by name pattern if provided
    if experiment_name_pattern:
        filtered_experiments = []
        for exp in all_experiments:
            exp_name = exp['name'].lstrip('/')
            if fnmatch.fnmatch(exp_name, experiment_name_pattern):
                filtered_experiments.append(exp)
        all_experiments = filtered_experiments

        if verbose:
            print(f"Filtered to {len(all_experiments)} experiments matching '{experiment_name_pattern}'")

    if verbose:
        if parent_run_pattern is not None:
            print(f"Searching for parent runs with pattern: {parent_run_pattern}")
        else:
            print("Searching for all parent runs")
        if tag_filter is not None:
            print(f"Filtering by tag: {tag_filter}")
        if hostname_filter is not None:
            print(f"Filtering by hostname: {hostname_filter}")
        print("="*80)

    # Parse tag filter if provided (format: "key=value")
    tag_key, tag_value = None, None
    if tag_filter:
        if '=' in tag_filter:
            tag_key, tag_value = tag_filter.split('=', 1)
            tag_key = tag_key.strip()
            tag_value = tag_value.strip()
        else:
            tag_key = tag_filter.strip()
            tag_value = 'true'  # Default to true if no value specified

    results_summary = {
        'processed': [],
        'failed': [],
        'skipped': []
    }

    # Group sweeps by configuration if merging
    config_sweeps = defaultdict(list) if merge_same_config else None

    # Search across all experiments
    for exp in all_experiments:
        exp_name = exp['name'].lstrip('/')
        exp_id = exp['experiment_id']

        try:
            # Find parent runs matching the pattern
            parent_runs = client.list_parent_runs(
                experiment_ids=[exp_id],
                run_name_pattern=parent_run_pattern
            )

            if not parent_runs:
                continue

            # Filter by tag if specified
            if tag_key:
                filtered_runs = []
                tag_value_lower = str(tag_value).lower()
                for run in parent_runs:
                    run_tags = client._extract_tags(run['data'])
                    run_tag_value = run_tags.get(tag_key, '')
                    if str(run_tag_value).lower() == tag_value_lower:
                        filtered_runs.append(run)
                parent_runs = filtered_runs

            if verbose:
                print(f"\nExperiment: {exp_name}")
                print(f"  Found {len(parent_runs)} parent run(s)")

            # Process each parent run
            for parent_run in parent_runs:
                parent_tags = client._extract_tags(parent_run['data'])
                parent_name = parent_tags.get('mlflow.runName', 'N/A')
                parent_run_id = parent_run['info']['run_id']
                parent_start_time = parent_run['info'].get('start_time', 0)
                timestamp_str = format_timestamp(parent_start_time)

                if verbose:
                    print(f"  Processing: {parent_name} (ID: {parent_run_id[:8]}..., created: {timestamp_str})")

                # Get sweep results using run ID instead of name
                child_runs = client.get_child_runs(parent_run_id)

                if len(child_runs) == 0:
                    if verbose:
                        print(f"    ⚠ No child runs found")
                    continue

                # Convert child runs to DataFrame
                df_sweep = client.runs_to_dataframe(child_runs)

                # Filter by hostname if specified
                if hostname_filter is not None:
                    # Try different possible column names for hostname
                    hostname_cols = [
                        'param_server/misc.hostname',
                        'param_server/misc_hostname',
                        'param_misc.hostname',
                        'param_misc_hostname',
                    ]
                    hostname_col = None
                    for col in hostname_cols:
                        if col in df_sweep.columns:
                            hostname_col = col
                            break
                    
                    if hostname_col is not None:
                        # Use exact match for hostname
                        mask = df_sweep[hostname_col].astype(str) == hostname_filter
                        df_sweep = df_sweep[mask]
                        if len(df_sweep) == 0:
                            if verbose:
                                print(f"    ⚠ No runs matching hostname '{hostname_filter}'")
                            continue
                        if verbose:
                            print(f"    ✓ Filtered to {len(df_sweep)} runs on hostname '{hostname_filter}'")
                    else:
                        if verbose:
                            # Show available columns containing 'host' for debugging
                            host_cols = [c for c in df_sweep.columns if 'host' in c.lower()]
                            if host_cols:
                                print(f"    ⚠ Hostname column not found. Available host-related columns: {host_cols}")
                            else:
                                print(f"    ⚠ No hostname column found, skipping this sweep")
                        continue

                # Add parent run info
                df_sweep['parent_run_id'] = parent_run_id
                df_sweep['parent_run_name'] = parent_name

                # Add experiment URL for each run
                # Format: https://workspace-url/#mlflow/experiments/{exp_id}/runs/{run_id}
                workspace_url = client.workspace_url
                df_sweep['experiment_url'] = df_sweep['run_id'].apply(
                    lambda run_id: f"{workspace_url}/#mlflow/experiments/{exp_id}/runs/{run_id}"
                )

                # Extract model name from parameters (before postprocessing)
                model_name = extract_model_name_from_params(df_sweep)

                # Fall back to experiment name if not found in params
                if model_name == 'unknown':
                    if verbose:
                        print(f"    ⚠ Could not extract model from params, using experiment name")
                    model_name = exp_name.replace('/', '_').replace('--', '-')

                # Normalize model name
                normalized_model = normalize_model_name(model_name)

                # Extract metadata
                docker_runtime = 'unknown'
                runtime_col = 'param_server/docker.runtime'
                if runtime_col in df_sweep.columns:
                    if not df_sweep[runtime_col].isna().all():
                        docker_runtime = str(df_sweep[runtime_col].iloc[0])

                docker_image = 'unknown'
                docker_col = 'param_server/docker.image'
                if docker_col in df_sweep.columns:
                    if not df_sweep[docker_col].isna().all():
                        docker_image = extract_docker_image(df_sweep[docker_col].iloc[0])

                # Extract GPU model - pass docker_image for fallback detection
                gpu_model = 'unknown'
                gpu_col = 'param_server/misc.nvidia_gpu_models'
                if gpu_col in df_sweep.columns:
                    if not df_sweep[gpu_col].isna().all():
                        gpu_model = parse_gpu_model(df_sweep[gpu_col].iloc[0], docker_runtime, docker_image)
                else:
                    gpu_model = parse_gpu_model(None, docker_runtime, docker_image)

                framework = extract_framework_from_docker_image(docker_image)
                failed_requests = extract_failed_requests(df_sweep)

                # Extract hostname if available
                hostname = 'unknown'
                hostname_cols = [
                    'param_server/misc.hostname',
                    'param_server/misc_hostname',
                    'param_misc.hostname',
                    'param_misc_hostname',
                ]
                for hcol in hostname_cols:
                    if hcol in df_sweep.columns and not df_sweep[hcol].isna().all():
                        hostname = str(df_sweep[hcol].iloc[0])
                        break

                # Postprocess results (experiment_url will be preserved)
                df_clean = postprocess_sweep_results(df_sweep)

                # Parse configuration (pass exp_name and df for fallback extraction)
                config = parse_experiment_config(parent_name, exp_name=exp_name, df=df_sweep)
                config['framework'] = framework

                # Create config key for grouping
                config_key = (normalized_model, gpu_model, config['framework'],
                             config['parallelism'], config['precision'])

                # Prepare metadata
                metadata = {
                    'gpu': gpu_model,
                    'docker_runtime': docker_runtime,
                    'framework': framework,
                    'parallelism': config['parallelism'],
                    'precision': config['precision'],
                    'docker_image': docker_image,
                    'hostname': hostname,
                    'model_name': model_name,
                    'normalized_model': normalized_model,
                    'timestamp': timestamp_str,
                    'parent_run': parent_name,
                    'parent_run_id': parent_run_id
                }

                sweep_data = {
                    'df_clean': df_clean,
                    'metadata': metadata,
                    'parent_name': parent_name,
                    'parent_run_id': parent_run_id,
                    'timestamp': timestamp_str,
                    'finished_count': (df_sweep['status'] == 'FINISHED').sum(),
                    'total_count': len(df_sweep),
                    'failed_requests': failed_requests,
                    'exp_name': exp_name
                }

                if merge_same_config:
                    # Group by configuration
                    config_sweeps[config_key].append(sweep_data)
                else:
                    # Save individual sweep immediately
                    _save_sweep(sweep_data, output_dir, results_summary, verbose)

        except Exception as e:
            if verbose:
                print(f"  ✗ Error processing {exp_name}: {str(e)}")
            results_summary['failed'].append({
                'experiment': exp_name,
                'error': str(e)
            })

    # Process merged sweeps if merging is enabled
    if merge_same_config:
        for config_key, sweeps in config_sweeps.items():
            _merge_and_save_sweeps(sweeps, config_key, output_dir, results_summary, verbose)

    return results_summary

def _validate_sweeps_compatible(sweeps: list, verbose: bool = True) -> tuple[bool, list[str]]:
    """
    Check if sweeps can be safely merged

    Returns:
        Tuple of (is_compatible, list_of_warnings)
    """
    if len(sweeps) <= 1:
        return True, []

    warnings = []

    # Check 1: Docker images must match exactly
    docker_images = [s['metadata']['docker_image'] for s in sweeps]
    unique_images = set(docker_images)
    if len(unique_images) > 1:
        warnings.append(f"Docker images don't match: {unique_images}")
        return False, warnings

    # Check 2: Framework versions should match
    frameworks = [s['metadata']['framework'] for s in sweeps]
    unique_frameworks = set(frameworks)
    if len(unique_frameworks) > 1:
        warnings.append(f"Frameworks don't match: {unique_frameworks}")
        return False, warnings

    # Check 3: GPU models should match
    gpus = [s['metadata']['gpu'] for s in sweeps]
    unique_gpus = set(gpus)
    if len(unique_gpus) > 1:
        warnings.append(f"GPU models don't match: {unique_gpus}")
        return False, warnings

    # Check 4: Check for overlapping configurations with different performance
    dedupe_cols = ['random_input_len', 'random_output_len', 'max_concurrency']
    metric_cols = ['median_ttft_ms', 'median_itl_ms', 'median_tpot_ms', 'output_throughput']

    # Build a dict of config -> metrics for each sweep
    config_metrics = {}
    for sweep_idx, sweep in enumerate(sweeps):
        df = sweep['df_clean']
        available_dedupe = [col for col in dedupe_cols if col in df.columns]
        available_metrics = [col for col in metric_cols if col in df.columns]

        if not available_dedupe or not available_metrics:
            continue

        for _, row in df.iterrows():
            # Create config key
            config_key = tuple(row[col] for col in available_dedupe if pd.notna(row[col]))

            if config_key in config_metrics:
                # Compare metrics
                prev_sweep_idx, prev_metrics = config_metrics[config_key]
                current_metrics = {col: row[col] for col in available_metrics if pd.notna(row[col])}

                # Check if metrics differ significantly (>5% relative difference)
                for metric in available_metrics:
                    if metric in prev_metrics and metric in current_metrics:
                        prev_val = prev_metrics[metric]
                        curr_val = current_metrics[metric]

                        if prev_val != 0:
                            rel_diff = abs(curr_val - prev_val) / abs(prev_val)
                            if rel_diff > 0.05:  # 5% threshold
                                warnings.append(
                                    f"Config {config_key}: {metric} differs by {rel_diff*100:.1f}% "
                                    f"between sweep {prev_sweep_idx+1} ({prev_val:.2f}) "
                                    f"and sweep {sweep_idx+1} ({curr_val:.2f})"
                                )
            else:
                # Store metrics for this config
                config_metrics[config_key] = (
                    sweep_idx,
                    {col: row[col] for col in available_metrics if pd.notna(row[col])}
                )

    # If we have performance warnings, this is a concern but not necessarily a blocker
    has_perf_warnings = len(warnings) > 0

    return not has_perf_warnings, warnings


def extract_framework_version(docker_image: str) -> str:
    """
    Extract framework version from docker image name

    Args:
        docker_image: Docker image (e.g., "vllm/vllm-openai:v0.6.4.post1", "rocm/vllm:rocm7.0.0_vllm_0.11.2_20251210")

    Returns:
        Framework version (e.g., "v0.6.4", "v0.11.2", or "unknown")
    """
    if not docker_image or pd.isna(docker_image):
        return 'unknown'

    docker_str = str(docker_image).lower()

    # Try to extract vLLM version
    # Pattern 1: vllm/vllm-openai:v0.6.4.post1
    vllm_pattern1 = r'vllm[/-].*?:v?(\d+\.\d+\.\d+)'
    match = re.search(vllm_pattern1, docker_str)
    if match:
        return f"v{match.group(1)}"

    # Pattern 2: rocm/vllm:rocm7.0.0_vllm_0.11.2_20251210
    vllm_pattern2 = r'vllm[_-](\d+\.\d+\.\d+)'
    match = re.search(vllm_pattern2, docker_str)
    if match:
        return f"v{match.group(1)}"

    # Try to extract SGLang version
    sglang_pattern = r'sglang.*?:v?(\d+\.\d+\.\d+)'
    match = re.search(sglang_pattern, docker_str)
    if match:
        return f"v{match.group(1)}"

    # Try to extract TensorRT-LLM version
    trt_pattern = r'tensorrt[_-]llm.*?:v?(\d+\.\d+\.\d+)'
    match = re.search(trt_pattern, docker_str)
    if match:
        return f"v{match.group(1)}"

    # Generic version pattern (last resort)
    generic_pattern = r':v?(\d+\.\d+\.\d+)'
    match = re.search(generic_pattern, docker_str)
    if match:
        return f"v{match.group(1)}"

    return 'unknown'


def _merge_and_save_sweeps(sweeps: list, config_key: tuple, output_dir: str,
                           results_summary: dict, verbose: bool):
    """Merge multiple sweeps with same config and save to CSV, with validation"""

    if len(sweeps) == 1:
        # Only one sweep, save normally
        _save_sweep(sweeps[0], output_dir, results_summary, verbose)
        return

    # Validate sweeps are compatible
    is_compatible, warnings = _validate_sweeps_compatible(sweeps, verbose)

    if not is_compatible:
        # Save separately with warnings
        if verbose:
            print(f"\n    ⚠️  CANNOT MERGE {len(sweeps)} sweeps - incompatible configurations:")
            for warning in warnings:
                print(f"       - {warning}")
            print(f"    Saving {len(sweeps)} sweeps separately...")

        for sweep in sweeps:
            _save_sweep(sweep, output_dir, results_summary, verbose)
            # Add warning to summary
            if results_summary['processed']:
                results_summary['processed'][-1]['merge_warning'] = '; '.join(warnings)
        return

    # Proceed with merge
    dfs_to_merge = []
    metadata_base = sweeps[0]['metadata'].copy()

    # Collect timestamps and parent runs
    timestamps = []
    parent_runs = []
    total_finished = 0
    total_runs = 0
    total_failed_requests = 0

    for sweep in sweeps:
        dfs_to_merge.append(sweep['df_clean'])
        timestamps.append(sweep['timestamp'])
        parent_runs.append(sweep['parent_name'])
        total_finished += sweep['finished_count']
        total_runs += sweep['total_count']
        total_failed_requests += sweep['failed_requests']

    # Merge dataframes
    df_merged = pd.concat(dfs_to_merge, ignore_index=True)

    # Remove duplicate rows (same random_input_len, random_output_len, max_concurrency)
    dedupe_cols = ['random_input_len', 'random_output_len', 'max_concurrency']
    available_dedupe_cols = [col for col in dedupe_cols if col in df_merged.columns]

    pre_dedupe_count = len(df_merged)
    if available_dedupe_cols:
        # Keep the most recent (last) occurrence
        df_merged = df_merged.drop_duplicates(subset=available_dedupe_cols, keep='last')
    dedupe_count = pre_dedupe_count - len(df_merged)

    # SORT AFTER MERGING - Convert to numeric and sort
    sort_columns = []
    for col_name in ['random_input_len', 'random_output_len', 'max_concurrency']:
        if col_name in df_merged.columns:
            try:
                # Ensure columns are numeric
                df_merged[col_name] = pd.to_numeric(df_merged[col_name], errors='coerce')
                sort_columns.append(col_name)
            except Exception as e:
                if verbose:
                    print(f"    Warning: Could not convert column '{col_name}' to numeric: {e}")

    # Sort the merged dataframe
    if sort_columns:
        try:
            df_merged = df_merged.sort_values(by=sort_columns, ascending=True)
            df_merged = df_merged.reset_index(drop=True)
        except Exception as e:
            if verbose:
                print(f"    Warning: Could not sort merged data by {sort_columns}: {e}")

    # Extract framework version
    framework_version = extract_framework_version(metadata_base['docker_image'])

    # Update metadata
    metadata_base['timestamp'] = f"merged_{len(sweeps)}_sweeps_{min(timestamps)}_to_{max(timestamps)}"
    metadata_base['parent_run'] = f"merged: {', '.join(parent_runs[:3])}" + (
        f" ... (+{len(parent_runs)-3} more)" if len(parent_runs) > 3 else ""
    )
    metadata_base['sweep_count'] = len(sweeps)
    metadata_base['docker_image'] = sweeps[0]['metadata']['docker_image']  # All should match
    metadata_base['framework_version'] = framework_version

    # Create model-specific subfolder
    model_output_dir = Path(output_dir) / metadata_base['normalized_model']
    model_output_dir.mkdir(parents=True, exist_ok=True)

    # Create filename with framework version
    filename = (f"{metadata_base['normalized_model']}-{metadata_base['gpu']}-"
               f"{metadata_base['framework']}-{framework_version}-"
               f"{metadata_base['parallelism']}-"
               f"{metadata_base['precision']}-merged_{len(sweeps)}_sweeps.csv")
    filepath = model_output_dir / filename

    # Export merged results
    export_sweep_results(
        df_merged,
        str(filepath),
        metadata=metadata_base,
        filter_finished=True,
        drop_status=True,
        verbose=False
    )

    # Track success
    results_summary['processed'].append({
        'experiment': sweeps[0]['exp_name'],
        'parent_run': f"MERGED: {len(sweeps)} sweeps",
        'timestamp': f"merged ({min(timestamps)} to {max(timestamps)})",
        'model': metadata_base['model_name'],
        'normalized_model': metadata_base['normalized_model'],
        'gpu': metadata_base['gpu'],
        'docker_runtime': metadata_base['docker_runtime'],
        'framework': metadata_base['framework'],
        'framework_version': framework_version,
        'parallelism': metadata_base['parallelism'],
        'precision': metadata_base['precision'],
        'docker_image': metadata_base['docker_image'],
        'hostname': metadata_base.get('hostname', 'unknown'),
        'runs': f"{total_finished}/{total_runs}",
        'failed_requests': total_failed_requests,
        'output_file': str(filepath.relative_to(output_dir)),
        'sweep_count': len(sweeps),
        'deduped_rows': dedupe_count,
        'merge_warning': 'None'
    })

    if verbose:
        print(f"\n    ✅ MERGED {len(sweeps)} sweeps (validated compatible):")
        print(f"       Docker: {metadata_base['docker_image']}")
        print(f"       Framework: {metadata_base['framework']} {framework_version}")
        for i, sweep in enumerate(sweeps, 1):
            print(f"       {i}. {sweep['parent_name']} ({sweep['timestamp']}) - {sweep['finished_count']}/{sweep['total_count']} runs")
        print(f"    ✓ Exported {len(df_merged)} unique configurations ({dedupe_count} duplicates removed)")
        print(f"      to: {filepath.relative_to(output_dir)}")
        print(f"    Total: {total_finished}/{total_runs} runs, {total_failed_requests} failed requests")

def _save_sweep(sweep_data: dict, output_dir: str, results_summary: dict, verbose: bool):
    """Save a single sweep to CSV"""
    metadata = sweep_data['metadata']
    df_clean = sweep_data['df_clean']

    # Extract framework version
    framework_version = extract_framework_version(metadata['docker_image'])
    metadata['framework_version'] = framework_version

    # Create model-specific subfolder
    model_output_dir = Path(output_dir) / metadata['normalized_model']
    model_output_dir.mkdir(parents=True, exist_ok=True)

    # Create filename with framework version
    filename = (f"{metadata['normalized_model']}-{metadata['gpu']}-"
               f"{metadata['framework']}-{framework_version}-"
               f"{metadata['parallelism']}-"
               f"{metadata['precision']}-{metadata['timestamp']}.csv")
    filepath = model_output_dir / filename

    # Export results
    export_sweep_results(
        df_clean,
        str(filepath),
        metadata=metadata,
        filter_finished=True,
        drop_status=True,
        verbose=False
    )

    # Track success
    results_summary['processed'].append({
        'experiment': sweep_data['exp_name'],
        'parent_run': sweep_data['parent_name'],
        'timestamp': sweep_data['timestamp'],
        'model': metadata['model_name'],
        'normalized_model': metadata['normalized_model'],
        'gpu': metadata['gpu'],
        'docker_runtime': metadata['docker_runtime'],
        'framework': metadata['framework'],
        'framework_version': framework_version,
        'parallelism': metadata['parallelism'],
        'precision': metadata['precision'],
        'docker_image': metadata['docker_image'],
        'hostname': metadata.get('hostname', 'unknown'),
        'runs': f"{sweep_data['finished_count']}/{sweep_data['total_count']}",
        'failed_requests': sweep_data['failed_requests'],
        'output_file': str(filepath.relative_to(output_dir)),
        'sweep_count': 1,
        'deduped_rows': 0,
        'merge_warning': 'N/A'
    })

    if verbose:
        print(f"    ✓ Exported {sweep_data['finished_count']}/{sweep_data['total_count']} runs to: {filepath.relative_to(output_dir)}")


def print_processing_summary(summary: dict):
    """Print summary of processing results"""
    print("\n" + "="*80)
    print("PROCESSING SUMMARY")
    print("="*80)

    print(f"\n✓ Successfully processed: {len(summary['processed'])}")
    if summary['processed']:
        df_processed = pd.DataFrame(summary['processed'])
        print("\nProcessed experiments:")
        with pd.option_context('display.max_columns', None, 'display.width', None):
            display(df_processed)

        # Show model groupings
        if 'normalized_model' in df_processed.columns:
            print("\nModel groupings:")
            model_counts = df_processed.groupby('normalized_model').size().reset_index(name='sweeps')
            display(model_counts)

        # Show unique configurations
        if 'docker_image' in df_processed.columns:
            print("\nDocker images used:")
            docker_images = df_processed.groupby('docker_image').size().reset_index(name='count')
            display(docker_images)

        # Show framework distribution
        if 'framework' in df_processed.columns:
            print("\nFramework distribution:")
            frameworks = df_processed.groupby('framework').size().reset_index(name='count')
            display(frameworks)

        # Show configuration combinations
        config_cols = ['framework', 'parallelism', 'precision']
        if all(col in df_processed.columns for col in config_cols):
            print("\nConfiguration combinations:")
            configs = df_processed.groupby(config_cols).size().reset_index(name='count')
            display(configs)

        # Show failed requests summary
        if 'failed_requests' in df_processed.columns:
            total_failed = df_processed['failed_requests'].sum()
            print(f"\nTotal failed requests across all sweeps: {total_failed}")
            if total_failed > 0:
                print("\nSweeps with failed requests:")
                failed_df = df_processed[df_processed['failed_requests'] > 0][
                    ['experiment', 'parent_run', 'failed_requests', 'runs']
                ]
                display(failed_df)

    if summary['skipped']:
        print(f"\n⚠ Skipped: {len(summary['skipped'])}")
        df_skipped = pd.DataFrame(summary['skipped'])
        display(df_skipped)

    if summary['failed']:
        print(f"\n✗ Failed: {len(summary['failed'])}")
        df_failed = pd.DataFrame(summary['failed'])
        display(df_failed)


def calculate_comparisons(df: pd.DataFrame, gpu1: str = 'MI355X', gpu2: str = 'B200') -> pd.DataFrame:
    """
    Calculate performance comparisons between GPUs in specific order

    Args:
        df: Merged dataframe from merge_configurations
        gpu1: First GPU name (AMD - MI355X)
        gpu2: Second GPU name (NVIDIA - B200)

    Returns:
        Dataframe with comparison columns added in order
    """
    df_compare = df.copy()

    # Metric to use for comparisons
    metric = 'total_token_throughput'

    # Find columns for each framework/GPU combo
    def find_column(gpu: str, framework: str) -> Optional[str]:
        for col in df.columns:
            if metric in col and gpu in col and framework in col.lower():
                return col
        return None

    # Get all framework columns
    gpu1_vllm = find_column(gpu1, 'vllm')
    gpu1_sglang = find_column(gpu1, 'sglang')
    gpu1_atom = find_column(gpu1, 'atom')

    gpu2_vllm = find_column(gpu2, 'vllm')
    gpu2_sglang = find_column(gpu2, 'sglang')
    gpu2_trt = find_column(gpu2, 'trt')

    # Helper function for safe division
    def safe_divide(numerator_col: Optional[str], denominator_col: Optional[str]) -> pd.Series:
        if numerator_col is None or denominator_col is None:
            return pd.Series([np.nan] * len(df), index=df.index)
        denom = df[denominator_col].replace(0, np.nan)
        return df[numerator_col] / denom

    # Use cleaner column names without "speedup"
    # 1. vLLM vs vLLM
    df_compare["vLLM vs vLLM"] = safe_divide(gpu1_vllm, gpu2_vllm)

    # 2. vLLM vs TRT
    df_compare["vLLM vs TRT"] = safe_divide(gpu1_vllm, gpu2_trt)

    # 3. SGLang vs SGLang
    df_compare["SGLang vs SGLang"] = safe_divide(gpu1_sglang, gpu2_sglang)

    # 4. SGLang vs TRT
    df_compare["SGLang vs TRT"] = safe_divide(gpu1_sglang, gpu2_trt)

    # 5. Best vs Best (based on geomean)
    best_gpu1_framework = find_best_framework_by_geomean(df, gpu1, metric)
    best_gpu2_framework = find_best_framework_by_geomean(df, gpu2, metric)

    gpu1_best_col = find_column(gpu1, best_gpu1_framework.lower())
    gpu2_best_col = find_column(gpu2, best_gpu2_framework.lower())

    df_compare["Best vs Best"] = safe_divide(gpu1_best_col, gpu2_best_col)

    # Add columns showing which framework is best
    df_compare[f"{gpu1} best framework"] = best_gpu1_framework if gpu1_best_col else 'N/A'
    df_compare[f"{gpu2} best framework"] = best_gpu2_framework if gpu2_best_col else 'N/A'

    return df_compare


def find_best_framework_by_geomean(df: pd.DataFrame, gpu: str, metric: str = 'total_token_throughput') -> str:
    """
    Find best framework for a GPU based on geometric mean of pairwise speedups

    Strategy:
    1. For each configuration (row), compute speedup ratios between frameworks
    2. Take geometric mean of those speedups across all valid configurations
    3. Select framework with highest average speedup vs others

    Args:
        df: Merged dataframe
        gpu: GPU name
        metric: Metric to use for comparison

    Returns:
        Best framework name (uppercase)
    """
    # Find all framework columns for this GPU
    # MI355X: vllm, sglang, atom
    # B200: vllm, sglang, trt
    framework_cols = {}

    if 'MI355X' in gpu.upper():
        frameworks_to_check = ['vllm', 'sglang', 'atom']
    else:  # B200
        frameworks_to_check = ['vllm', 'sglang', 'trt']

    for col in df.columns:
        if gpu in col and metric in col:
            # Extract framework
            for fw in frameworks_to_check:
                if fw in col.lower():
                    framework_cols[fw.upper()] = col
                    break

    if not framework_cols:
        return 'unknown'

    if len(framework_cols) == 1:
        return list(framework_cols.keys())[0]

    # Calculate average speedup for each framework vs all others
    framework_scores = {}

    print(f"\n  DEBUG {gpu} framework comparison (row-by-row speedups):")

    for fw1, col1 in framework_cols.items():
        speedups = []

        # Compare fw1 against all other frameworks
        for fw2, col2 in framework_cols.items():
            if fw1 == fw2:
                continue

            # Compute row-by-row speedup: fw1 / fw2
            # Only include rows where both frameworks have valid values
            mask = (df[col1] > 0) & (df[col2] > 0) & df[col1].notna() & df[col2].notna()
            if mask.sum() > 0:
                row_speedups = (df.loc[mask, col1] / df.loc[mask, col2]).values
                speedups.extend(row_speedups)

        # Calculate geometric mean of all speedups for this framework
        if len(speedups) > 0:
            geomean = np.exp(np.log(speedups).mean())
            framework_scores[fw1] = geomean
            print(f"    {fw1}: geomean speedup vs others = {geomean:.3f}x (from {len(speedups)} comparisons)")
        else:
            framework_scores[fw1] = 0
            print(f"    {fw1}: no valid comparisons")

    # Select framework with highest score
    if framework_scores:
        best_framework = max(framework_scores.items(), key=lambda x: x[1])[0]
        best_score = framework_scores[best_framework]
        print(f"  -> Best: {best_framework} with average speedup={best_score:.3f}x")
        return best_framework

    return 'unknown'

def load_model_results(model_dir: Path, track_additions: bool = False) -> Dict[tuple, pd.DataFrame]:
    """
    Load all CSV results for a single model

    Args:
        model_dir: Path to model directory containing CSV files
        track_additions: If True, also return list of all loaded files with timestamps

    Returns:
        Dictionary mapping (gpu, framework, version, precision, parallelism) -> DataFrame
        If same key exists multiple times, keeps the one with latest timestamp
        
        If track_additions=True, returns tuple: (results_dict, additions_list)
    """
    results = {}
    results_timestamps = {}  # Track timestamps to keep latest
    all_additions = []  # Track all loaded files for "recent additions" feature
    csv_files = list(model_dir.glob("*.csv"))

    for csv_file in csv_files:
        try:
            df = pd.read_csv(csv_file)

            # Extract metadata from first row
            if len(df) > 0 and 'gpu' in df.columns and 'framework' in df.columns:
                gpu = df['gpu'].iloc[0]
                framework = df['framework'].iloc[0]

                # Get framework version
                if 'framework_version' in df.columns:
                    version = df['framework_version'].iloc[0]
                elif 'docker_image' in df.columns:
                    version = extract_framework_version(df['docker_image'].iloc[0])
                else:
                    version = 'unknown'

                # Get precision and parallelism to differentiate configs
                precision = df['precision'].iloc[0] if 'precision' in df.columns else 'unknown'
                parallelism = df['parallelism'].iloc[0] if 'parallelism' in df.columns else 'unknown'

                # Get timestamp from CSV or filename
                timestamp = '00000000-000000'
                if 'timestamp' in df.columns and len(df) > 0:
                    ts_val = df['timestamp'].iloc[0]
                    if pd.notna(ts_val) and str(ts_val) != 'nan':
                        ts_str = str(ts_val)
                        # Handle merged timestamp format: "merged_N_sweeps_YYYYMMDD-HHMMSS_to_YYYYMMDD-HHMMSS"
                        merged_match = re.search(r'merged_\d+_sweeps_(\d{8}-\d{6})_to_(\d{8}-\d{6})', ts_str)
                        if merged_match:
                            # Use the end date of the merge range
                            timestamp = merged_match.group(2)
                        elif re.match(r'^\d{8}-\d{6}$', ts_str):
                            # Standard timestamp format
                            timestamp = ts_str
                        else:
                            # Try to find any timestamp pattern in the string
                            any_ts_match = re.search(r'(\d{8}-\d{6})', ts_str)
                            if any_ts_match:
                                timestamp = any_ts_match.group(1)
                
                # If still default, try to extract from filename
                if timestamp == '00000000-000000':
                    ts_match = re.search(r'(\d{8}-\d{6})', csv_file.name)
                    if ts_match:
                        timestamp = ts_match.group(1)
                    elif 'merged' in csv_file.name.lower():
                        # For merged files without timestamp, use file modification time
                        try:
                            mtime = os.path.getmtime(csv_file)
                            timestamp = datetime.fromtimestamp(mtime).strftime('%Y%m%d-%H%M%S')
                        except OSError:
                            timestamp = 'merged'

                key = (gpu, framework, version, precision, parallelism)
                
                # Track this file for additions list
                if track_additions:
                    model_name = 'unknown'
                    if 'model_name' in df.columns and len(df) > 0:
                        model_name = df['model_name'].iloc[0]
                    elif 'model' in df.columns and len(df) > 0:
                        model_name = df['model'].iloc[0]
                    
                    docker_image = 'unknown'
                    if 'docker_image' in df.columns and len(df) > 0:
                        docker_image = df['docker_image'].iloc[0]
                    
                    all_additions.append({
                        'timestamp': timestamp,
                        'model': model_name,
                        'gpu': gpu,
                        'framework': framework,
                        'version': version,
                        'precision': precision,
                        'parallelism': parallelism,
                        'docker_image': docker_image,
                        'filename': csv_file.name,
                        'num_configs': len(df)
                    })
                
                # Keep the latest timestamp version if key already exists
                if key in results:
                    old_timestamp = results_timestamps.get(key, '')
                    if timestamp > old_timestamp:
                        # Replace with newer file
                        results[key] = df
                        results_timestamps[key] = timestamp
                    # else: keep existing (older file, skip this one)
                else:
                    results[key] = df
                    results_timestamps[key] = timestamp

        except Exception as e:
            print(f"Warning: Could not load {csv_file.name}: {e}")

    if track_additions:
        return results, all_additions
    return results


def parse_version_for_sorting(version_str: str) -> tuple:
    """
    Parse version string for proper sorting.
    Returns tuple that sorts correctly: (major, minor, patch, suffix)
    
    Examples:
        "v0.15.0" -> (0, 15, 0, "")
        "v0.9.0" -> (0, 9, 0, "")
        "unknown" -> (-1, 0, 0, "unknown")
    """
    if not version_str or version_str == 'unknown':
        return (-1, 0, 0, version_str or "unknown")
    
    # Try to parse version like "v0.15.0" or "0.15.0"
    match = re.match(r'v?(\d+)\.(\d+)\.?(\d*)(.*)', str(version_str))
    if match:
        major = int(match.group(1))
        minor = int(match.group(2))
        patch = int(match.group(3)) if match.group(3) else 0
        suffix = match.group(4) or ""
        return (major, minor, patch, suffix)
    
    # Fallback: return string for sorting
    return (-1, 0, 0, str(version_str))


def select_latest_version(results: Dict[tuple, pd.DataFrame]) -> Tuple[Dict[tuple, Tuple[str, str, pd.DataFrame]], List[str]]:
    """
    For each (gpu, framework, precision, parallelism) combination, select the latest version

    Args:
        results: Dictionary mapping (gpu, framework, version, precision, parallelism) -> DataFrame

    Returns:
        Tuple of (Dictionary mapping (gpu, framework, precision, parallelism) -> (version, docker_image, DataFrame), warnings)
    """
    # Group by (gpu, framework, precision, parallelism)
    grouped = {}
    for (gpu, framework, version, precision, parallelism), df in results.items():
        key = (gpu, framework, precision, parallelism)
        if key not in grouped:
            grouped[key] = []

        # Extract docker image from first row
        docker_image = 'unknown'
        if 'docker_image' in df.columns and len(df) > 0:
            docker_image = df['docker_image'].iloc[0]
        
        # Extract timestamp for secondary sorting (if versions are equal)
        timestamp = '00000000-000000'
        if 'timestamp' in df.columns and len(df) > 0:
            timestamp = str(df['timestamp'].iloc[0])

        grouped[key].append((version, docker_image, df, timestamp))

    # Select latest version for each group
    selected = {}
    warnings = []

    for (gpu, framework, precision, parallelism), versions in grouped.items():
        if len(versions) > 1:
            # Sort by parsed version (major, minor, patch), then by timestamp
            versions_sorted = sorted(
                versions, 
                key=lambda x: (parse_version_for_sorting(x[0]), x[3]),  # (version_tuple, timestamp)
                reverse=True
            )
            latest_version, latest_docker, latest_df, _ = versions_sorted[0]

            warnings.append(
                f"Multiple versions found for {gpu} {framework} {precision} {parallelism}: "
                f"{[v for v, _, _, _ in versions]}. Using latest: {latest_version}"
            )
        else:
            latest_version, latest_docker, latest_df, _ = versions[0]

        selected[(gpu, framework, precision, parallelism)] = (latest_version, latest_docker, latest_df)

    return selected, warnings


def group_results_by_config(selected_results: Dict[tuple, Tuple[str, str, pd.DataFrame]]) -> Dict[tuple, Dict[tuple, Tuple[str, str, pd.DataFrame]]]:
    """
    Group results by (precision, parallelism) configuration

    Args:
        selected_results: Dict mapping (gpu, framework, precision, parallelism) -> (version, docker_image, DataFrame)

    Returns:
        Dict mapping (precision, parallelism) -> {(gpu, framework) -> (version, docker_image, DataFrame)}
    """
    config_groups = {}

    for (gpu, framework, precision, parallelism), (version, docker_image, df) in selected_results.items():
        config_key = (precision, parallelism)

        if config_key not in config_groups:
            config_groups[config_key] = {}

        config_groups[config_key][(gpu, framework)] = (version, docker_image, df)

    return config_groups


def _create_short_sheet_name(model_name: str, precision: str, parallelism: str,
                              used_names: set, max_len: int = 31) -> str:
    """
    Create a short, readable Excel sheet name

    Args:
        model_name: Model name (e.g., "Llama-3.3-70B", "Qwen3-235B")
        precision: Precision (e.g., "fp8", "bf16")
        parallelism: Parallelism config (e.g., "tp1", "tp8")
        used_names: Set of already used sheet names
        max_len: Maximum sheet name length (Excel limit is 31)

    Returns:
        Unique sheet name within character limit
    """
    # Model name abbreviations (base names without variant suffixes)
    model_abbrevs = {
        'Llama-3.3-70B': 'Llama3.3-70B',
        'Llama-3.1-405B': 'Llama3.1-405B',
        'Llama-3.1-70B': 'Llama3.1-70B',
        'Llama-3.1-8B': 'Llama3.1-8B',
        'Qwen3-235B': 'Qwen3-235B',
        'Qwen2.5-72B': 'Qwen2.5-72B',
        'GLM-4.7': 'GLM-4.7',
        'DeepSeek-R1': 'DSR1',
        'DeepSeek-V3.1': 'DSV3.1',
        'DeepSeek-V3': 'DSV3',
        'Mistral-Large': 'Mistral-L',
        'Mixtral-8x22B': 'Mixtral-8x22',
        'Mixtral-8x7B': 'Mixtral-8x7',
        'Llama-4-Maverick-17B-128E': 'Llama4-Mav',
    }
    
    # Track if original model name is an Instruct variant
    is_instruct = 'Instruct' in model_name or 'instruct' in model_name

    # Normalize model name for lookup (strip MoE details, version dates, etc.)
    lookup_name = model_name
    
    # Handle Qwen3 MoE models: "Qwen3-235B-A22B-Instruct-2507" -> "Qwen3-235B"
    qwen_moe_match = re.match(r'(Qwen\d*-\d+B)-A\d+B.*', model_name)
    if qwen_moe_match:
        lookup_name = qwen_moe_match.group(1)
    
    # Handle Llama 4 MoE models: "Llama-4-Maverick-17B-128E-Instruct" -> "Llama-4-Maverick-17B-128E"
    llama4_match = re.match(r'(Llama-4-Maverick-\d+B-\d+E).*', model_name)
    if llama4_match:
        lookup_name = llama4_match.group(1)
    
    # Handle DeepSeek models with version dates: "DeepSeek-R1-0528" -> "DeepSeek-R1"
    deepseek_match = re.match(r'(DeepSeek-[RV]\d+(?:\.\d+)?)-\d+', model_name)
    if deepseek_match:
        lookup_name = deepseek_match.group(1)

    # Use abbreviation if available, otherwise use original
    short_model = model_abbrevs.get(lookup_name, model_name)
    
    # If still not found, try partial matching for common model families
    if short_model == model_name and len(model_name) > 15:
        for key, abbrev in model_abbrevs.items():
            if model_name.startswith(key):
                short_model = abbrev
                break
    
    # Append "-I" suffix for Instruct variants (if not already present)
    if is_instruct and not short_model.endswith('-I'):
        short_model = short_model + '-I'

    # Further shorten if still too long
    if len(short_model) > 15:
        # Take first letter of each word/part
        parts = short_model.replace('-', ' ').replace('.', ' ').split()
        if len(parts) > 2:
            # Keep first part and abbreviate rest
            short_model = parts[0][:6] + '-' + ''.join(p[0] for p in parts[1:])

    # Format precision (skip if unknown)
    precision_fmt = precision.upper() if precision.lower() != 'unknown' else None

    # Format parallelism (uppercase)
    parallelism_fmt = parallelism.upper() if parallelism.lower() != 'unknown' else None

    # Build sheet name: "Model Precision Parallelism"
    # e.g., "Llama3.3-70B FP8 TP1"
    # Skip unknown values to keep names clean
    parts = [short_model]
    if precision_fmt:
        parts.append(precision_fmt)
    if parallelism_fmt:
        parts.append(parallelism_fmt)
    sheet_name = ' '.join(parts)

    # Truncate if needed
    if len(sheet_name) > max_len:
        # Try removing spaces
        sheet_name = f"{short_model}_{precision_fmt}_{parallelism_fmt}"[:max_len]

    # Ensure uniqueness
    if sheet_name not in used_names:
        return sheet_name

    # Add numeric suffix if duplicate
    counter = 2
    while True:
        suffix = f"_{counter}"
        candidate = sheet_name[:max_len - len(suffix)] + suffix
        if candidate not in used_names:
            return candidate
        counter += 1

def create_comparative_analysis(
    parent_dir: str,
    output_file: str = "comparative_analysis.xlsx",
    gpu1: str = 'MI355X',
    gpu2: str = 'B200',
    verbose: bool = True
) -> Dict[str, pd.DataFrame]:
    """
    Create comparative analysis across all models, separating by precision/parallelism
    """
    parent_path = Path(parent_dir)

    if not parent_path.exists():
        raise ValueError(f"Directory not found: {parent_dir}")

    model_dirs = [d for d in parent_path.iterdir() if d.is_dir()]

    if verbose:
        print(f"Found {len(model_dirs)} model directories")
        print("="*80)

    all_results = {}
    all_warnings = []
    summary_data = []

    for model_dir in model_dirs:
        model_name = model_dir.name

        if verbose:
            print(f"\nProcessing model: {model_name}")

        raw_results = load_model_results(model_dir)

        if not raw_results:
            if verbose:
                print(f"  ⚠ No valid CSV files found")
            continue

        if verbose:
            print(f"  Found {len(raw_results)} result files")

        selected_results, warnings = select_latest_version(raw_results)

        if warnings:
            all_warnings.extend([f"{model_name}: {w}" for w in warnings])
            if verbose:
                for warning in warnings:
                    print(f"  ⚠ {warning}")

        # Group by (precision, parallelism) configuration
        config_groups = group_results_by_config(selected_results)

        if verbose:
            print(f"  Found {len(config_groups)} configuration(s): {list(config_groups.keys())}")

        # Process each configuration separately
        for (precision, parallelism), config_results in config_groups.items():
            # Create unique sheet name with config
            config_suffix = f"{parallelism}_{precision}"
            sheet_key = f"{model_name}_{config_suffix}"

            if verbose:
                print(f"\n  Processing config: {precision} / {parallelism}")
                print(f"    GPU/Framework combinations: {list(config_results.keys())}")

            merged_df = merge_configurations(config_results, gpu1=gpu1, gpu2=gpu2)

            if merged_df.empty:
                if verbose:
                    print(f"    ⚠ No data after merging")
                continue

            comparison_df = calculate_comparisons(merged_df, gpu1, gpu2)

            summary_row = {
                'Model': model_name,
                'Precision': precision,
                'Parallelism': parallelism,
            }

            # Add best frameworks to summary
            if f"{gpu1} best framework" in comparison_df.columns:
                summary_row[f'{gpu1} Best'] = comparison_df[f"{gpu1} best framework"].iloc[0]
            if f"{gpu2} best framework" in comparison_df.columns:
                summary_row[f'{gpu2} Best'] = comparison_df[f"{gpu2} best framework"].iloc[0]

            # Add comparison geomeans
            comparison_cols = ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]
            for comp_col in comparison_cols:
                if comp_col in comparison_df.columns:
                    geomean = calculate_geomean_speedup(comparison_df, comp_col)
                    summary_row[comp_col] = geomean if not np.isnan(geomean) else np.nan

            summary_data.append(summary_row)

            if verbose:
                print(f"\n    Summary:")
                print(f"      Configurations: {len(comparison_df)}")

                if f"{gpu1} best framework" in comparison_df.columns:
                    best_gpu1 = comparison_df[f"{gpu1} best framework"].iloc[0]
                    best_gpu2 = comparison_df[f"{gpu2} best framework"].iloc[0]
                    print(f"      {gpu1} best framework (geomean): {best_gpu1}")
                    print(f"      {gpu2} best framework (geomean): {best_gpu2}")

                for comp_col in comparison_cols:
                    if comp_col in comparison_df.columns:
                        geomean = calculate_geomean_speedup(comparison_df, comp_col)
                        print(f"      {comp_col}: {geomean:.3f}x (geomean)")

            all_results[sheet_key] = comparison_df

    summary_df = pd.DataFrame(summary_data)

    if not summary_df.empty:
        # Reorder summary columns
        base_cols = ['Model', 'Precision', 'Parallelism', f'{gpu1} Best', f'{gpu2} Best']
        comparison_cols = ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]

        available_cols = [col for col in base_cols + comparison_cols if col in summary_df.columns]
        summary_df = summary_df[available_cols]

        # Sort by Model, Parallelism, Precision for better organization
        sort_cols = [col for col in ['Model', 'Parallelism', 'Precision'] if col in summary_df.columns]
        if sort_cols:
            summary_df = summary_df.sort_values(by=sort_cols).reset_index(drop=True)

    if all_results:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)

                # Format summary sheet with merged comparison headers
                ws_summary = writer.sheets['Summary']
                format_summary_sheet(ws_summary, summary_df, gpu1, gpu2)

            # Track used sheet names to avoid duplicates
            used_sheet_names = {'Summary'}

            for sheet_key, df in all_results.items():
                # Parse sheet_key to extract model, parallelism, precision
                # Format is: "model_name_parallelism_precision"
                parts = sheet_key.rsplit('_', 2)
                if len(parts) >= 3:
                    model_part = parts[0]
                    parallelism_part = parts[1]
                    precision_part = parts[2]
                else:
                    # Fallback
                    model_part = sheet_key
                    parallelism_part = 'unknown'
                    precision_part = 'unknown'

                # Create short, readable sheet name
                sheet_name = _create_short_sheet_name(
                    model_part, precision_part, parallelism_part, used_sheet_names
                )
                used_sheet_names.add(sheet_name)

                df_with_summary = add_geomean_summary(df, gpu1, gpu2)
                df_with_summary.to_excel(writer, sheet_name=sheet_name, index=False)
                format_sheet_with_merged_headers(writer.sheets[sheet_name], df_with_summary, gpu1, gpu2)

        if verbose:
            print("\n" + "="*80)
            print(f"✓ Exported comparative analysis to: {output_file}")
            print(f"  Summary sheet with {len(summary_df)} model configurations")
            print(f"  Individual sheets: {len(all_results)}")

    if all_warnings:
        print("\n" + "="*80)
        print("WARNINGS:")
        for warning in all_warnings:
            print(f"  ⚠ {warning}")

    return all_results

def merge_configurations(dataframes: Dict[tuple, Tuple[str, str, pd.DataFrame]],
                        gpu1: str = 'MI355X',
                        gpu2: str = 'B200') -> pd.DataFrame:
    """
    Merge multiple dataframes by configuration with ordered columns

    Args:
        dataframes: Dict mapping (gpu, framework) -> (version, docker_image, DataFrame)
        gpu1: First GPU (AMD - MI355X)
        gpu2: Second GPU (NVIDIA - B200)

    Returns:
        Merged dataframe with sorted columns
    """
    if not dataframes:
        return pd.DataFrame()

    # Get all unique configurations
    all_configs = set()
    for (gpu, framework), (version, docker_image, df) in dataframes.items():
        df_norm = normalize_dataframe(df)
        for _, row in df_norm.iterrows():
            config = (
                row.get('random_input_len'),
                row.get('random_output_len'),
                row.get('max_concurrency')
            )
            if all(pd.notna(v) for v in config):
                all_configs.add(config)

    # Sort configurations
    all_configs = sorted(list(all_configs))

    # Build merged dataframe
    config_df = pd.DataFrame(all_configs, columns=['random_input_len', 'random_output_len', 'max_concurrency'])

    metric_columns = ['median_itl_ms', 'median_ttft_ms', 'median_tpot_ms',
                     'median_e2el_ms', 'output_throughput', 'total_token_throughput']

    # Define raw data column order:
    # MI355X: VLLM, SGLANG, ATOM
    # B200: VLLM, SGLANG, TRT
    column_order = [
        (gpu1, 'vllm'),
        (gpu1, 'sglang'),
        (gpu1, 'atom'),
        (gpu2, 'vllm'),
        (gpu2, 'sglang'),
        (gpu2, 'trt'),
    ]

    # Add data in specified order
    for gpu, framework in column_order:
        key = (gpu, framework)

        if key in dataframes:
            version, docker_image, df = dataframes[key]
            df_norm = normalize_dataframe(df)

            # Use full docker image in column header
            col_prefix = f"{gpu} {framework.upper()} ({docker_image})"

            # Add metrics
            for metric in metric_columns:
                if metric in df_norm.columns:
                    # Create mapping from config to value
                    config_to_value = {}
                    for _, row in df_norm.iterrows():
                        config = (
                            row.get('random_input_len'),
                            row.get('random_output_len'),
                            row.get('max_concurrency')
                        )
                        if all(pd.notna(v) for v in config):
                            config_to_value[config] = row[metric]

                    # Add column
                    col_name = f"{col_prefix}_{metric}"
                    config_df[col_name] = config_df.apply(
                        lambda row: config_to_value.get(
                            (row['random_input_len'], row['random_output_len'], row['max_concurrency']),
                            np.nan
                        ),
                        axis=1
                    )
        else:
            # Framework missing for this GPU - add NaN columns
            col_prefix = f"{gpu} {framework.upper()} (N/A)"
            for metric in metric_columns:
                col_name = f"{col_prefix}_{metric}"
                config_df[col_name] = np.nan

    return config_df


def extract_metadata_from_results(selected_results: Dict[tuple, Tuple[str, str, pd.DataFrame]]) -> Dict[str, str]:
    """
    Extract common metadata (precision, parallelism) from selected results

    Args:
        selected_results: Dict mapping (gpu, framework) -> (version, docker_image, DataFrame)

    Returns:
        Dictionary with metadata fields
    """
    metadata = {
        'precision': 'N/A',
        'parallelism': 'N/A'
    }

    # Get first available dataframe to extract metadata
    for (gpu, framework), (version, docker_image, df) in selected_results.items():
        if len(df) > 0:
            # Extract precision
            if 'precision' in df.columns:
                metadata['precision'] = df['precision'].iloc[0]

            # Extract parallelism
            if 'parallelism' in df.columns:
                metadata['parallelism'] = df['parallelism'].iloc[0]

            break

    return metadata

def add_geomean_summary(df: pd.DataFrame, gpu1: str = 'MI355X', gpu2: str = 'B200') -> pd.DataFrame:
    """
    Add summary rows with geomean for each comparison column

    Args:
        df: Comparison dataframe
        gpu1: First GPU name
        gpu2: Second GPU name

    Returns:
        Dataframe with summary rows appended
    """
    # Find comparison columns (without "speedup" in the name now)
    comparison_cols = ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]

    # Filter to columns that actually exist in the dataframe
    existing_comparison_cols = [col for col in comparison_cols if col in df.columns]

    if not existing_comparison_cols:
        return df

    # Create summary rows
    summary_rows = []

    # Empty separator row
    separator = pd.Series([''] * len(df.columns), index=df.columns)
    summary_rows.append(separator)

    # Header row
    header = pd.Series([''] * len(df.columns), index=df.columns)
    header['random_input_len'] = '=== GEOMEAN SUMMARY ==='
    summary_rows.append(header)

    # Geomean row - put ALL geomeans in one row
    geomean_row = pd.Series([''] * len(df.columns), index=df.columns)
    geomean_row['random_input_len'] = 'Geomean'

    for comp_col in existing_comparison_cols:
        geomean = calculate_geomean_speedup(df, comp_col)
        # Store as numeric value (no "x" suffix)
        geomean_row[comp_col] = geomean if not np.isnan(geomean) else ''

    summary_rows.append(geomean_row)

    # Append summary rows
    df_summary = pd.concat([df] + [pd.DataFrame([row]) for row in summary_rows], ignore_index=True)

    return df_summary


def normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize dataframe to have consistent columns and data types

    Args:
        df: Input dataframe

    Returns:
        Normalized dataframe with key metrics
    """
    key_columns = [
        'random_input_len', 'random_output_len', 'max_concurrency',
        'median_itl_ms', 'median_ttft_ms', 'median_tpot_ms',
        'median_e2el_ms', 'output_throughput', 'total_token_throughput'
    ]

    # Select only available columns
    available_cols = [col for col in key_columns if col in df.columns]
    df_norm = df[available_cols].copy()

    # Convert numeric columns
    for col in ['random_input_len', 'random_output_len', 'max_concurrency']:
        if col in df_norm.columns:
            df_norm[col] = pd.to_numeric(df_norm[col], errors='coerce')

    # Sort by configuration
    sort_cols = [col for col in ['random_input_len', 'random_output_len', 'max_concurrency']
                 if col in df_norm.columns]
    if sort_cols:
        df_norm = df_norm.sort_values(by=sort_cols).reset_index(drop=True)

    return df_norm



def extract_framework_from_column(col_name: str) -> str:
    """Extract framework name from column name"""
    for fw in ['vllm', 'sglang', 'atom', 'trt']:
        if fw in col_name.lower():
            return fw.upper()
    return 'unknown'


def calculate_geomean_speedup(df: pd.DataFrame, comparison_col: str) -> float:
    """
    Calculate geometric mean of speedup values, handling inf and nan

    Args:
        df: Dataframe with comparison column
        comparison_col: Name of comparison column

    Returns:
        Geometric mean speedup
    """
    if comparison_col not in df.columns:
        return np.nan

    values = df[comparison_col].replace([np.inf, -np.inf], np.nan).dropna()

    if len(values) == 0:
        return np.nan

    # Filter out any remaining invalid values
    values = values[values > 0]

    if len(values) == 0:
        return np.nan

    # Geometric mean of speedup
    return np.exp(np.log(values).mean())

def print_comparison_summary(results: Dict[str, pd.DataFrame],
                            gpu1: str = 'MI355X',
                            gpu2: str = 'B200'):
    """
    Print summary of comparisons across all models

    Args:
        results: Dictionary from create_comparative_analysis
        gpu1: First GPU name
        gpu2: Second GPU name
    """
    print("\n" + "="*80)
    print("COMPARATIVE ANALYSIS SUMMARY")
    print("="*80)

    summary_data = []

    for model_name, df in results.items():
        row = {'model': model_name}

        # Show best frameworks first
        if f"{gpu1} best framework" in df.columns:
            row[f'{gpu1}_best'] = df[f"{gpu1} best framework"].iloc[0]
        if f"{gpu2} best framework" in df.columns:
            row[f'{gpu2}_best'] = df[f"{gpu2} best framework"].iloc[0]

        # Find speedup columns in order
        speedup_order = [
            f"{gpu1} vs {gpu2} VLLM speedup",
            f"{gpu1} VLLM vs {gpu2} TRT speedup",
            f"{gpu1} vs {gpu2} SGLANG speedup",
            f"{gpu1} SGLANG vs {gpu2} TRT speedup",
            f"{gpu1} vs {gpu2} Best speedup"
        ]

        for col in speedup_order:
            if col in df.columns:
                geomean = calculate_geomean_speedup(df, col)
                # Simplify column name for display
                col_display = col.replace(f"{gpu1} vs {gpu2} ", "").replace(f"{gpu1} ", "").replace(f"{gpu2} ", "").replace(" speedup", "")
                row[col_display] = f"{geomean:.3f}x" if not np.isnan(geomean) else "N/A"

        summary_data.append(row)

    summary_df = pd.DataFrame(summary_data)
    display(summary_df)



def format_sheet_with_merged_headers(worksheet, df: pd.DataFrame, gpu1: str = 'MI355X', gpu2: str = 'B200'):
    """
    Format Excel sheet with merged headers for GPU/Framework groups and comparisons
    """
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    metric_columns = ['median_itl_ms', 'median_ttft_ms', 'median_tpot_ms',
                     'median_e2el_ms', 'output_throughput', 'total_token_throughput']

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    missing_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')  # Light red

    # Insert blank row at row 1
    worksheet.insert_rows(1)

    # Helper function to extract framework header (everything before the metric)
    def extract_framework_header(col_name: str) -> str:
        """Extract 'GPU FRAMEWORK (DOCKER)' from 'GPU FRAMEWORK (DOCKER)_metric_name'"""
        for metric in metric_columns:
            if col_name.endswith(f'_{metric}'):
                return col_name[:-len(f'_{metric}')]
        return col_name

    # Build a mapping of framework headers to their column indices
    framework_to_cols = {}
    for col_idx, col_name in enumerate(df.columns, 1):
        is_metric_col = any(col_name.endswith(f'_{metric}') for metric in metric_columns)
        if is_metric_col:
            framework_header = extract_framework_header(col_name)
            if framework_header not in framework_to_cols:
                framework_to_cols[framework_header] = []
            framework_to_cols[framework_header].append(col_idx)

    # Parse DataFrame column names to identify groups
    framework_groups = []
    comparison_groups = []
    col_idx = 1

    for col_name in df.columns:
        if col_name in ['random_input_len', 'random_output_len', 'max_concurrency']:
            col_idx += 1
            continue

        comparison_cols = ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]
        if col_name in comparison_cols:
            if not comparison_groups or comparison_groups[-1]['end_col'] != col_idx - 1:
                comparison_groups.append({'start_col': col_idx, 'end_col': col_idx})
            else:
                comparison_groups[-1]['end_col'] = col_idx
            col_idx += 1
            continue

        if 'best framework' in col_name:
            col_idx += 1
            continue

        is_metric_col = any(col_name.endswith(f'_{metric}') for metric in metric_columns)

        if is_metric_col:
            framework_header = extract_framework_header(col_name)
            if not framework_groups or framework_groups[-1]['header'] != framework_header:
                framework_groups.append({'header': framework_header, 'start_col': col_idx, 'end_col': col_idx})
            else:
                framework_groups[-1]['end_col'] = col_idx

        col_idx += 1

    # Process each column header
    col_idx = 1
    comparison_col_indices = []
    metric_col_indices = []

    for col_name in df.columns:
        col_letter = get_column_letter(col_idx)

        if col_name in ['random_input_len', 'random_output_len', 'max_concurrency']:
            worksheet.merge_cells(f'{col_letter}1:{col_letter}2')
            cell = worksheet[f'{col_letter}1']
            cell.value = col_name
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True, size=9)
            cell.fill = header_fill
            cell.border = thin_border

        elif 'best framework' in col_name:
            worksheet.merge_cells(f'{col_letter}1:{col_letter}2')
            cell = worksheet[f'{col_letter}1']
            cell.value = col_name
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, size=8)
            cell.fill = header_fill
            cell.border = thin_border

        elif col_name in ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]:
            cell_row2 = worksheet.cell(row=2, column=col_idx)
            cell_row2.value = col_name
            cell_row2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_row2.font = Font(bold=True, size=8)
            cell_row2.fill = header_fill
            cell_row2.border = thin_border
            comparison_col_indices.append(col_idx)

        else:
            cell_row2 = worksheet.cell(row=2, column=col_idx)
            for metric in metric_columns:
                if col_name.endswith(f'_{metric}'):
                    cell_row2.value = metric
                    metric_col_indices.append(col_idx)
                    break
            cell_row2.alignment = Alignment(horizontal='center', vertical='center')
            cell_row2.font = Font(bold=True, size=8)
            cell_row2.fill = header_fill
            cell_row2.border = thin_border

        col_idx += 1

    # Write framework headers to row 1 and merge
    for group in framework_groups:
        start_col_letter = get_column_letter(group['start_col'])
        end_col_letter = get_column_letter(group['end_col'])

        if group['start_col'] != group['end_col']:
            worksheet.merge_cells(f'{start_col_letter}1:{end_col_letter}1')

        cell = worksheet[f'{start_col_letter}1']
        cell.value = group['header']
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.border = thin_border

        for col in range(group['start_col'], group['end_col'] + 1):
            worksheet.cell(row=1, column=col).border = thin_border
            worksheet.cell(row=1, column=col).fill = header_fill

    # Write comparison header
    for group in comparison_groups:
        start_col_letter = get_column_letter(group['start_col'])
        end_col_letter = get_column_letter(group['end_col'])

        if group['start_col'] != group['end_col']:
            worksheet.merge_cells(f'{start_col_letter}1:{end_col_letter}1')

        cell = worksheet[f'{start_col_letter}1']
        cell.value = f"{gpu1} vs {gpu2}"
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.border = thin_border

        for col in range(group['start_col'], group['end_col'] + 1):
            worksheet.cell(row=1, column=col).border = thin_border
            worksheet.cell(row=1, column=col).fill = header_fill

    # Format all data cells (starting from row 3)
    for row_idx in range(3, worksheet.max_row + 1):
        # Check if this is a summary/geomean row
        first_cell_value = worksheet.cell(row=row_idx, column=1).value
        is_summary_row = first_cell_value in ['=== GEOMEAN SUMMARY ===', 'Geomean', '']
        is_geomean_row = first_cell_value == 'Geomean'

        # For each framework group, check if all metric values are zero or missing
        framework_missing = {}
        for fw_header, fw_cols in framework_to_cols.items():
            all_zero_or_missing = True
            for fw_col_idx in fw_cols:
                cell_val = worksheet.cell(row=row_idx, column=fw_col_idx).value
                if cell_val is not None and cell_val != '' and cell_val != 0:
                    try:
                        if float(cell_val) != 0:
                            all_zero_or_missing = False
                            break
                    except (ValueError, TypeError):
                        all_zero_or_missing = False
                        break
            framework_missing[fw_header] = all_zero_or_missing and not is_summary_row

        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Bold the geomean row
            if is_geomean_row:
                cell.font = Font(bold=True)

            # Only apply missing data highlighting to metric columns (not config columns)
            if col_idx in metric_col_indices and not is_summary_row:
                # Find which framework this column belongs to
                col_framework = None
                for fw_header, fw_cols in framework_to_cols.items():
                    if col_idx in fw_cols:
                        col_framework = fw_header
                        break

                # Check if this framework has all zeros/missing for this row
                if col_framework and framework_missing.get(col_framework, False):
                    cell.fill = missing_fill
                # Also check individual cell for missing/zero
                elif cell.value is None or cell.value == '' or (isinstance(cell.value, (int, float)) and (pd.isna(cell.value) or cell.value == 0)):
                    cell.fill = missing_fill

            # Format and highlight comparison columns
            if col_idx in comparison_col_indices:
                if not is_summary_row:
                    # Check for missing data in comparison columns
                    if cell.value is None or cell.value == '' or (isinstance(cell.value, float) and pd.isna(cell.value)):
                        cell.fill = missing_fill
                    elif isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                        cell.number_format = '0.0%'
                elif is_geomean_row and isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                    # Format geomean row comparison values as percentage
                    cell.number_format = '0.0%'

    # Column widths - tighter
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        col_name = df.columns[col_idx - 1]

        if col_name in ['random_input_len', 'random_output_len', 'max_concurrency']:
            worksheet.column_dimensions[col_letter].width = 10
        elif 'vs' in col_name or 'best framework' in col_name:
            worksheet.column_dimensions[col_letter].width = 12
        else:
            worksheet.column_dimensions[col_letter].width = 12

    worksheet.freeze_panes = 'D3'

def format_summary_sheet(worksheet, df: pd.DataFrame, gpu1: str = 'MI355X', gpu2: str = 'B200'):
    """
    Format summary sheet with merged headers and heatmap coloring for comparison columns
    """
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

    # Heatmap colors
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')      # >= 100% (light green)
    light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # 85-100% (lighter green)
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')      # 70-85% (yellow)
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')         # < 70% (red)

    # Insert row at top
    worksheet.insert_rows(1)

    # Find comparison columns
    comparison_cols = ["vLLM vs vLLM", "vLLM vs TRT", "SGLang vs SGLang", "SGLang vs TRT", "Best vs Best"]
    comparison_start = None
    comparison_end = None
    comparison_col_indices = []

    for col_idx, col_name in enumerate(df.columns, 1):
        if col_name in comparison_cols:
            if comparison_start is None:
                comparison_start = col_idx
            comparison_end = col_idx
            comparison_col_indices.append(col_idx)

    # Format all columns
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)

        if col_name in comparison_cols:
            cell_row2 = worksheet.cell(row=2, column=col_idx)
            cell_row2.value = col_name
            cell_row2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_row2.font = Font(bold=True, size=8)
            cell_row2.fill = header_fill
            cell_row2.border = thin_border
        else:
            worksheet.merge_cells(f'{col_letter}1:{col_letter}2')
            cell = worksheet[f'{col_letter}1']
            cell.value = col_name
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, size=9)
            cell.fill = header_fill
            cell.border = thin_border

    # Merge comparison header
    if comparison_start and comparison_end:
        start_letter = get_column_letter(comparison_start)
        end_letter = get_column_letter(comparison_end)
        worksheet.merge_cells(f'{start_letter}1:{end_letter}1')

        cell = worksheet[f'{start_letter}1']
        cell.value = f"{gpu1} vs {gpu2}"
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size=9)
        cell.fill = header_fill
        cell.border = thin_border

        for col_idx in range(comparison_start, comparison_end + 1):
            worksheet.cell(row=1, column=col_idx).border = thin_border
            worksheet.cell(row=1, column=col_idx).fill = header_fill

    # Format all data cells (starting from row 3) with heatmap coloring
    for row_idx in range(3, worksheet.max_row + 1):
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply heatmap coloring to comparison columns
            if col_idx in comparison_col_indices:
                if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                    cell.number_format = '0.0%'
                    value = cell.value  # Already a ratio (e.g., 0.85 = 85%)

                    if value >= 1.0:  # >= 100%
                        cell.fill = green_fill
                    elif value >= 0.85:  # 85-100%
                        cell.fill = light_green_fill
                    elif value >= 0.70:  # 70-85%
                        cell.fill = yellow_fill
                    else:  # < 70%
                        cell.fill = red_fill
                # Leave blank/None cells uncolored

    # Set column widths - tighter
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        col_name = df.columns[col_idx - 1]

        if col_name == 'Model':
            worksheet.column_dimensions[col_letter].width = 14
        elif col_name in ['Precision', 'Parallelism']:
            worksheet.column_dimensions[col_letter].width = 10
        elif 'Best' in col_name:
            worksheet.column_dimensions[col_letter].width = 10
        else:
            worksheet.column_dimensions[col_letter].width = 11

    worksheet.freeze_panes = 'A3'
